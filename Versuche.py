from openpyxl import load_workbook
import matplotlib.pyplot as plt
global test
test = 0


def getdata(wanted):
    wb = load_workbook("/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPG_Test.xlsx", data_only=True)
    data = []
    years = []
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "1954-1963":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row = y+1, column = 1)
                if a1.value == wanted:
                    for x in range(0, 20, 2):
                        b1 = wb.active.cell(row=y+1, column=x + 2)
                        data.append(b1.value/1000)

        elif wb.sheetnames[s] == "1964-1966":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row = y+1, column = 1)
                if a1.value == wanted:
                    for x in range(0, 4, 2):
                        b1 = wb.active.cell(row=y+1, column=x + 2)
                        data.append(b1.value/1000)
                    for x in range(4, 6, 2):
                        data.append(b1.value)

        elif wb.sheetnames[s] == "1967":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row = y+1, column = 1)
                if a1.value == wanted:
                    for x in range(0, 2, 2):
                        b1 = wb.active.cell(row=y+1, column=x + 2)
                        data.append(b1.value)

        elif wb.sheetnames[s] == "1968-1972":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row=y + 1, column=1)
                if a1.value == wanted:
                    for x in range(0, 10, 2):
                        b1 = wb.active.cell(row=y + 1, column=x + 2)
                        data.append(b1.value)

        elif wb.sheetnames[s] == "1973-1986":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row=y + 1, column=1)
                if a1.value == wanted:
                    for x in range(0, 28, 2):
                        b1 = wb.active.cell(row=y + 1, column=x + 2)
                        data.append(b1.value)

        elif wb.sheetnames[s] == "1987-1997":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row=y + 1, column=1)
                if a1.value == wanted:
                    for x in range(0, 22, 2):
                        b1 = wb.active.cell(row=y + 1, column=x + 2)
                        data.append(b1.value)

        elif wb.sheetnames[s] == "1998-2002":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row=y + 1, column=1)
                if a1.value == wanted:
                    for x in range(0, 10, 2):
                        b1 = wb.active.cell(row=y + 1, column=x + 2)
                        data.append(b1.value)

    for y in range(1954, 2003):
        years.append(y)


    print(data)

    plt.plot(years, data, "m")
    plt.title('Haushaltsplan')
    plt.ylabel("Eingaben/Ausgaben")
    plt.xlabel("Jahre")
    plt.text(60, .025, "Hallo")
    plt.show()

wanted = "Test"
getdata(wanted)

