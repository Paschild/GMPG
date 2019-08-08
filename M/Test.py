from openpyxl import load_workbook
import matplotlib.pyplot as plt
import wx
import wx.grid
from collections import defaultdict

global istodersoll
istodersoll = 1
global istodersoll_text
istodersoll_text = ""


class Fenster(wx.Frame):
    def __init__(self, parent, title):
        '''Constructor, der das Fenster erzeugt'''
        # wx.Frame.__init__(self, parent=None, title = "Fenster", size=(1000, 700))
        super(Fenster, self).__init__(parent, title=title, size=(800, 800))

        panel = wx.Panel(self, size=(800, 700))
        panel.SetBackgroundColour("gray")
        box = wx.BoxSizer(wx.VERTICAL)

        self.label = wx.StaticText(panel, label="Übersicht", style=wx.ALIGN_CENTER)
        box.Add(self.label, 0, wx.EXPAND | wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 15)

        self.button = wx.Button(panel, 0, label="Close", )
        self.button.Bind(wx.EVT_BUTTON, self.bei_Click)

        self.myGrid = MyGrid(panel)       # ruft Grid-Klasse
        box.Add(self.myGrid)

        panel.SetSizer(box)
        self.Centre()
        self.Show()

    def bei_Click(self, event):
        self.Close()

class MyGrid(wx.grid.Grid):                 # verwendet Template-Dictionary new_dict_template
    def __init__(self, parent):
        wx.grid.Grid.__init__(self, parent)

        self.parent = parent
        self.CreateGrid(135, 10)            # erzeugt Grid und legt Größe fest
        self.EnableEditing(False)           # deaktiviert Bearbeitung
        self.Show()


        '''Überschriften'''
        spalte = 0
        for key in new_dict_template:
            self.SetColLabelValue(spalte, key)
            self.SetCellFont(0, spalte, wx.Font(16, wx.ROMAN, wx.NORMAL, wx.BOLD))
            self.SetCellBackgroundColour(0, spalte, wx.LIGHT_GREY)
            spalte = spalte + 1

        '''Zeilen füllen'''
        y = 0
        for key in new_dict_template:
            x = 2
            for value in new_dict_template[key]:
                self.SetCellValue(x, y, value)
                self.SetCellFont(x, y, wx.Font(14, wx.ROMAN, wx.NORMAL, wx.NORMAL))
                x = x + 1
            y = y + 1

        for i in range(10):                 # legt Größe der Spalten fest
            self.SetColSize(i, 240)
        self.SetRowLabelSize(0)             # versteckt Label-Zeile ganz links

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self)
        self.SetSizer(sizer)

        self.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.select_cell)

    def select_cell(self, event):
        print(new_template_cells[(event.GetCol()+1, event.GetRow())].value)
        print(event.GetCol(), event.GetRow())
        print(new_template_cells)
        pass



class Template_Cell:
    def __init__(self, row, column, value, color, sheet):
        self.row = row
        self.column = column
        self.value = value
        self.color = color
        self.sheet = sheet
        #self.konzepte = konzepte           #später kommt noch self.konzepte dazu


class Cell:
    def __init__(self, row, column, value, year, sheet, istodersoll):
        self.row = row
        self.column = column
        self.value = value
        self.year = year
        self.sheet = sheet
        if istodersoll == 1:
            self.istodersoll = "IST"
        elif istodersoll == 2:
            self.istodersoll = "SOLL"

    def ausgabe(self):
        print("Mappe: " + str(self.sheet))
        print("Jahr: " + str(self.year))
        print(str(self.istodersoll) + "-Wert")
        print("Spalte: " +str(self.column))
        print("Zeile: " + str(self.row))
        print("Wert: " + str(self.value))


def gettemplate():
    file_location = "/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPI_Template.xlsx"
    wb = load_workbook(file_location, data_only=True)
    template_cells = {}     # key = (zeile, spalt); value = Template_Cell obj
    dict_template = defaultdict(list)
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] != "Synthese":
            wb.active = sheet
            spalte = 1
            for zeile in range(10, 200):
                inhalt = wb.active.cell(row = zeile, column = spalte)
                if not inhalt.value:
                    continue
                else:
                    template_cells[(spalte, zeile)] = Template_Cell(zeile, spalte, inhalt.value, getattr(inhalt.fill.fgColor, inhalt.fill.fgColor.type), wb.sheetnames[sheet])
                    dict_template[wb.sheetnames[sheet]].append(inhalt.value)
    return template_cells, dict_template


def getdata(wanted):
    wb = load_workbook(filename, data_only=True)
    global istodersoll
    data = []
    years = []
    cells = []
    g = gen_years()
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] == "1954-1963":
            wb.active = sheet
            for zeile in range(1, 100):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 20, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column= istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1964-1966":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 4, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))
                    for spalte in range(4, 6, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1967":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 2, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1968-1972":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1973-1986":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 28, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1987-1997":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 22, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1998-2002":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

    return cells


def gen_years():
    for x in range(1954, 2003):
        yield x


def plot_matplot(new_cells):
    years = []
    data = []
    for cell in new_cells:
        years.append(cell.year)
        data.append(cell.value)

    plt.plot(years, data, "m")
    plt.title("Haushaltsplan" + istodersoll_text)
    plt.ylabel("Einnahmen/Ausgaben")
    plt.xlabel("Jahre")
    plt.text(60, .025, "Hallo")
    plt.show()


filename = "/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPG_Test.xlsx"
wanted = "Test"


#new_cells = getdata(wanted)
#plot_matplot(new_cells)

new_template_cells, new_dict_template = gettemplate()
#for cell in new_template_cells:
    #if cell.sheet == "1998-2002":
        #print(cell.value)


#for cell in new_cells:
    #cell.ausgabe()
    #break

app = wx.App()
print(new_dict_template)
print(new_template_cells)
frame = Fenster(None, "Fenster")
frame.Show()
app.MainLoop()




