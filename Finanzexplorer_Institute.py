# -*- coding: utf-8 -*-
import wx
import wx.grid
import os
import os.path
import Finanzexplorer_model as model
import copy
import json
import math
from collections import defaultdict, OrderedDict
from operator import itemgetter
from openpyxl import load_workbook
from matplotlib import use
use('WXAgg')
import matplotlib.pyplot as plt
from fpdf import FPDF
import datetime
import locale
locale.setlocale(locale.LC_ALL, 'de_DE')

'''
Eine Übersicht über alle wx.Python widgets findet sich unter folgendem Link:
https://wxpython.org/Phoenix/docs/html/gallery.html
'''

' Git-Test'

develope_mode = False
WHITE = "#ffffff"           # für das Interface (wx.grid)
active_konzeptColor = "#ffffff"     # active bedeutet, dieses Konzept ist angewählt
active_konzept = None

worksheets = ["1954-1963", "1964-1966", "1967", "1968-1972", "1973-1986", "1987-1997", "1998-2002"]

model.RECHNUNGSTYP = ""         # siehe Finanzexplorer_model.py. Ursprünglich zur Unterscheidung von 'EA' und 'VÜ'
                                # EA = Einnahmen/Ausgabenrechnung; VÜ = Vermögensübersicht
                                # Mit den Instituten kam 'INST' hinzu.

dct_cells = {}
found_cells = []    # für die Suchfunktion


class PDF(FPDF):
    def header(self):
        self.set_font("Times", "B", 10)
        # self.image("Icon.png", 10, 8, 8)
        self.cell(80)
        self.cell(30, 10, frame.GetTitle(), 0, 1, 'C')
        self.ln(10)

    def footer(self):
        today = datetime.date.today()
        self.set_y(-15)
        self.set_font("Times", 'I', 8)
        self.cell(0, 10, str(today), 0, 0, 'C')

        self.set_y(-15)
        link = self.add_link()
        self.set_link(link, page=1)
        self.cell(0, 10, "top", 0, 0, 'R', link=link)


class Konzept:
    def __init__(self, name, color):
        self.rechnungstyp = model.RECHNUNGSTYP
        self.name = name
        self.color = color

        self.cells = []     # not the real cell_objects, just the position as (row, col);
                            # Beim laden eines gespeicherten Konzepts werden diese Zellen in der Konzeptfarbe markiert

        self.plots = defaultdict(list)
        self.uncertain = []


class Cell:
    def __init__(self, cellrow, cellcol, value):
        self.row = cellrow
        self.col = cellcol
        self.value = value
        self.color = WHITE
        self.konzept = None


class ExcelCell:
    def __init__(self, row=None, col=None, path=None, sheet=None, year=None, typ=None, category=None, betrag=None):
        self.row = row
        self.col = col
        self.path = path
        self.sheet = sheet
        self.year = year
        self.type = typ         # z.B.: IST oder SOLL
        self.category = category     # z.B.: Gesamteinnahmen    (noch str, besser FK)

        # ab diesem Zeitpunkt sind die Werte in den Haushaltsplänen in TDM angegeben
        if (self.year > 1965 and self.type == "IST") or (self.year > 1966 and self.type == "SOLL"):
            self.betrag = betrag
        else:
            self.betrag = betrag / 1000

        # Die Einnahmen sind ab 1997/98 negativ aufgeführt. Dies hatten wir so in den Excel-Tabellen übernommen.
        if ((self.year > 1997 and self.type == "IST") or (self.year > 1998 and self.type == "SOLL")) and self.row < 25:
            self.betrag = self.betrag * -1

        if self.year <= 1999:
            self.inflationsbereinigt = self.betrag / dct_preisindices[self.year]
        else:
            # # noch Falsch. Was passiert mit den Werten, für die wir keinen Verbraucherpreisindex haben? (nach 1999)
            # self.inflationsbereinigt = self.betrag
            # # habe es nun in line_plot_inst() geändert. Inflationsbereinigt werden die Daten nach 1999 nicht angezeigt
            pass


# Das erste Fenster, welches sich beim starten des Programms öffnet.
class DialogRechnungstypInit(wx.Dialog):
    def __init__(self, parent, title):
        super(DialogRechnungstypInit, self).__init__(parent, title=title, size=(280, 160))

        # ---- Widgets erzeugen
        self.panel = wx.Panel(self)
        self.radiobox = wx.RadioBox(self.panel, -1, label="",  # https://wxpython.org/Phoenix/docs/html/wx.RadioBox.html
                                    choices=["Institutes",
                                             "Einnahmen-/Ausgabenrechnung (EA)",
                                             "Vermögensübersicht (VÜ)",
                                             "saved Concepts"],
                                    majorDimension=0,
                                    style=wx.RA_SPECIFY_ROWS)

        self.button = wx.Button(self.panel, -1, label="Okay")

        # --- Widgets positionieren mit wx.BoxSizer (https://wxpython.org/Phoenix/docs/html/sizers_overview.html)
        self.topsizer = wx.BoxSizer(wx.VERTICAL)
        self.topsizer.AddSpacer(2)
        self.topsizer.Add(self.radiobox, flag=wx.ALIGN_CENTER)
        self.topsizer.AddSpacer(8)
        self.topsizer.Add(self.button, flag=wx.ALIGN_CENTER)

        self.panel.SetSizer(self.topsizer)

        # --- Widgets an Functionen binden (Hier: wenn 'okay'-Button gedrückt wird, wird self.start ausgeführt)
        # --- https://wxpython.org/Phoenix/docs/html/events_overview.html
        self.button.Bind(wx.EVT_BUTTON, self.start)

        # Erst wenn dieses Fenster geschlossen wurde, kann mit Hauptfenster gearbeitet werden.
        # Alternative zu Modal: .Show()
        self.ShowModal()

    def start(self, _):
        tmp_rechnungstyp_id = self.radiobox.GetSelection()
        if tmp_rechnungstyp_id == 0:
            model.RECHNUNGSTYP = "INST"
            frame.set_interface(_, "INST")
            self.Close()
        elif tmp_rechnungstyp_id == 1:
            model.RECHNUNGSTYP = "EA"
            frame.set_interface(_, "EA")
            self.Close()
        elif tmp_rechnungstyp_id == 2:
            model.RECHNUNGSTYP = "VÜ"
            frame.set_interface(_, "VÜ")
            self.Close()
        elif tmp_rechnungstyp_id == 3:
            if frame.load_konzepte_picker(_):
                self.Close()
        else:
            pass


# Einstellungsfenster, welches sich öffnet, bevor die Grafen für die Gesamtrechnung angezeigt werden
class DialogGesamtPlotSettings(wx.Dialog):
    def __init__(self, parent, title, lst):
        super(DialogGesamtPlotSettings, self).__init__(parent, title=title, size=(250, 130))

        self.parent = parent
        self.panel = wx.Panel(self)

        self.radiobutton_sum = wx.RadioButton(self.panel, -1, label="show Sum of institutes per file", style=wx.RB_GROUP)
        self.radiobutton_inst = wx.RadioButton(self.panel, -1, label="show institutes seperate")
        self.radiobutton_all = wx.RadioButton(self.panel, -1, label="show Sum and Institutes seperate")

        self.btn = wx.Button(self.panel, -1, label="okay")

        # --- Positionierung
        self.sizer_top = wx.BoxSizer(wx.VERTICAL)
        self.sizer_top.Add(self.radiobutton_sum)
        self.sizer_top.Add(self.radiobutton_inst)
        self.sizer_top.Add(self.radiobutton_all)
        self.sizer_top.Add(self.btn)

        self.panel.SetSizer(self.sizer_top)

        self.btn.Bind(wx.EVT_BUTTON, self.create_plot)

    def create_plot(self, _):
        if self.radiobutton_sum.GetValue():
            mode = 1
        elif self.radiobutton_inst.GetValue():
            mode = 2
        else:
            mode = 3
        line_plot_gesamt(frame.new_get_konzept(), mode)
        self.Close()


class DialogPlotSettings(wx.Dialog):
    def __init__(self, parent, title):
        super(DialogPlotSettings, self).__init__(parent, title=title, size=(250, 130))

        self.parent = parent
        self.panel = wx.Panel(self)

        self.text01 = wx.StaticText(self.panel, -1, label="Ist- oder Sollwerte: ")
        self.choiceTyp = wx.Choice(self.panel, -1, choices=["IST", "SOLL", "IST & SOLL"], size=(100, 20))
        self.text02 = wx.StaticText(self.panel, -1, label="Ein Diagramm pro.. ")
        self.choiceGrouping = wx.Choice(self.panel, -1, choices=["Institut", "Konzept"], size=(100, 20))
        self.text03 = wx.StaticText(self.panel, -1, label="max Y-Wert: ")
        self.choiceYTicks = wx.Choice(self.panel, -1, choices=["auto", "compareable"], size=(100, 20))

        self.checkbox_inflation = wx.CheckBox(self.panel, -1, label="Inflationsbereinigt")

        self.btn = wx.Button(self.panel, -1, label="okay")

        self.sizer_settings = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer_left = wx.BoxSizer(wx.VERTICAL)
        self.sizer_right = wx.BoxSizer(wx.VERTICAL)

        self.sizer_left.AddSpacer(4)
        self.sizer_left.Add(self.text01)
        self.sizer_left.AddSpacer(8)
        self.sizer_left.Add(self.text02)
        self.sizer_left.AddSpacer(8)
        self.sizer_left.Add(self.text03)
        self.sizer_left.AddSpacer(10)
        self.sizer_left.Add(self.checkbox_inflation)

        self.sizer_right.AddSpacer(4)
        self.sizer_right.Add(self.choiceTyp)
        self.sizer_right.AddSpacer(5)
        self.sizer_right.Add(self.choiceGrouping)
        self.sizer_right.AddSpacer(5)
        self.sizer_right.Add(self.choiceYTicks)
        self.sizer_right.AddSpacer(10)
        self.sizer_right.Add(self.btn)

        self.sizer_settings.Add(self.sizer_left, flag=wx.ALIGN_TOP)
        self.sizer_settings.AddSpacer(5)
        self.sizer_settings.Add(self.sizer_right, flag=wx.ALIGN_TOP)
        self.sizer_settings.AddSpacer(5)

        self.panel.SetSizer(self.sizer_settings)

        self.btn.Bind(wx.EVT_BUTTON, self.create_plot)

    def create_plot(self, _):
        checkbox_inflation = 1 if self.checkbox_inflation.GetValue() else 0
        # --- EndModal um einen Integer zurückzugeben. Damit ein Integer mehrere Informationen beinhalten kann,
        # habe ich mit den Dezimalstellen gearbeitet.
        # Erscheint mir nicht optimal und ist sicherlich eleganter zu lösen ;)
        self.EndModal((self.choiceTyp.GetSelection()+1) * 10000 +
                      (self.choiceGrouping.GetSelection()+1) * 1000 +
                      (self.choiceYTicks.GetSelection()+1) * 100 +
                      (checkbox_inflation+1) * 10 +
                      9)                    # Prüfnummer (Damit kein Plot kommt, falls das Fenster geschlossen wird)


class DialogNewKonzept(wx.Dialog):
    def __init__(self, parent, title):
        super(DialogNewKonzept, self).__init__(parent, title=title, size=(200, 100))
        self.parent = parent
        self.panel = wx.Panel(self)

        self.name_of_new_konzept = wx.TextCtrl(self.panel, -1, "", size=(100, 23))
        self.color_of_new_konzept = wx.ColourPickerCtrl(self.panel, -1, wx.RED)

        self.btn = wx.Button(self.panel, -1, label="okay")

        self.sizer_dialog = wx.BoxSizer(wx.VERTICAL)
        self.sizer_dialog_konzept = wx.BoxSizer(wx.HORIZONTAL)

        self.sizer_dialog_konzept.Add(self.name_of_new_konzept, flag=wx.ALIGN_CENTER)
        self.sizer_dialog_konzept.AddSpacer(10)
        self.sizer_dialog_konzept.Add(self.color_of_new_konzept, flag=wx.ALIGN_CENTER)

        self.sizer_dialog.AddSpacer(5)
        self.sizer_dialog.Add(self.sizer_dialog_konzept, flag=wx.ALIGN_CENTER)
        self.sizer_dialog.AddSpacer(10)
        self.sizer_dialog.Add(self.btn, flag=wx.ALIGN_CENTER)

        self.panel.SetSizer(self.sizer_dialog)

        self.btn.Bind(wx.EVT_BUTTON, self.create_new_konzept)

    def create_new_konzept(self, _):
        self.parent.add_new_konzept(self.name_of_new_konzept.Value, rgb_to_hex(self.color_of_new_konzept.GetColour()))
        # rgb_to_hex() ist unten definiert. keine build-in Funktion von Python
        self.Close()


class MyGrid(wx.grid.Grid): # 'Mouse vs. Python' hat mir Anfangs sehr geholfen, um mit den Grids in wxPython umzugehen.
                            # http://www.blog.pythonlibrary.org/2010/04/04/wxpython-grid-tips-and-tricks/
    def __init__(self, parent):     # Durch die erste __init__ wird nichts mehr vom parent (wx.grid.Grid) geerbt..
        """Constructor"""
        wx.grid.Grid.__init__(self, parent)     # ..um dennoch zu erben, muss die parent.__init__ aufgerufen werden
        self.parent = parent
        self.CreateGrid(200, 58)
        # self.SetRowSize(0, 60)
        for i in range(58):
            self.SetColSize(i, 240)

        self.SetRowLabelSize(0)
        self.SetMargins(0, 0)

        self.this_row = 0
        self.this_col = 0

        self.GetGridWindow().Bind(wx.EVT_MOTION, self.on_mouse_over)

        self.Show()

    def on_mouse_over(self, event):
        """
        Displays a tooltip over any cell in a certain column
        """
        # This method was suggested by none other than Robin Dunn
        # http://www.blog.pythonlibrary.org/2010/04/04/wxpython-grid-tips-and-tricks/
        # https://alldunn.com/robin/
        x, y = self.CalcUnscrolledPosition(event.GetX(), event.GetY())
        coords = self.XYToCell(x, y)
        try:
            if model.RECHNUNGSTYP != "INST":
                msg = "{} {}: {}".format(model.get_dct_cells()[(coords[0], coords[1])].value[0].bezeichnung,
                                         model.get_dct_cells()[(coords[0], coords[1])].jahr,
                                         model.get_dct_cells()[(coords[0], coords[1])].posten.geldbetrag)
                event.GetEventObject().SetToolTip(msg)
            else:
                pass
        except KeyError:  # for empty cells
            pass
        except AttributeError:  # for cells without oberkategorie
            pass

    def show_popup_menu(self, event):
        self.this_row = event.GetRow()
        self.this_col = event.GetCol()
        if not hasattr(self, "popupID1"):
            self.popupID1 = wx.NewId()
            self.popupID2 = wx.NewId()
            self.popupID3 = wx.NewId()
            self.popupID4 = wx.NewId()
            self.popupID5 = wx.NewId()
            self.popupID6 = wx.NewId()

        menu = wx.Menu()
        item = wx.MenuItem(menu, self.popupID1, "trigger Cell")
        item_02 = wx.MenuItem(menu, self.popupID2, "trigger category")
        item_03 = wx.MenuItem(menu, self.popupID3, "select whole category")
        item_04 = wx.MenuItem(menu, self.popupID4, "unselect whole Category")
        item_uncertain_cat = wx.MenuItem(menu, self.popupID5, "uncertain")
        item_uncertain = wx.MenuItem(menu, self.popupID5, "uncertain")

        sub_menu = wx.Menu()
        sub_menu.Append(item_02)
        sub_menu.Append(item_03)
        sub_menu.Append(item_04)
        sub_menu.Append(item_uncertain_cat)

        menu.Append(item)
        menu.Append(wx.NewId(), "Category", sub_menu)
        menu.Append(item_uncertain)

        self.PopupMenu(menu)

        self.Bind(wx.EVT_MENU, self.select_cell, item)
        self.Bind(wx.EVT_MENU, self.trigger_kategorie_from_popup, item_02)
        self.Bind(wx.EVT_MENU, self.select_whole_category_from_popup, item_03)
        self.Bind(wx.EVT_MENU, self.unselect_whole_category_from_popup, item_04)
        self.Bind(wx.EVT_MENU, self.select_uncertain_from_popup, item_uncertain)
        self.Bind(wx.EVT_MENU, self.select_uncertain_cat_from_popup, item_uncertain_cat)
        menu.Destroy()

    def trigger_kategorie_from_popup(self, _):
        kat_id = self.get_cell(self.this_row, self.this_col).value[0].id
        for c in model.get_dct_cells().values():
            if c.value[0].id == kat_id:
                if c.color != active_konzeptColor:
                    c.color = active_konzeptColor
                    c.konzept = active_konzept
                    if (c.row, c.col) not in self.parent.GetParent().konzepte[c.konzept.name].cells:
                        self.parent.GetParent().konzepte[c.konzept.name].cells.append((c.row, c.col))
                else:
                    if c.color != WHITE:
                        self.parent.GetParent().konzepte[c.konzept.name].cells.remove((c.row, c.col))
                    c.color = WHITE
                    c.konzept = None
                self.SetCellBackgroundColour(c.row, c.col, c.color)
        self.ForceRefresh()

    def select_whole_category_from_popup(self, _):
        kat_id = self.get_cell(self.this_row, self.this_col).value[0].id
        for c in model.get_dct_cells().values():
            if c.value[0].id == kat_id:
                c.color = active_konzeptColor
                c.konzept = active_konzept
                if (c.row, c.col) not in self.parent.GetParent().konzepte[c.konzept.name].cells:
                    self.parent.GetParent().konzepte[c.konzept.name].cells.append((c.row, c.col))
                self.SetCellBackgroundColour(c.row, c.col, c.color)
        self.ForceRefresh()

    def unselect_whole_category_from_popup(self, _):
        kat_id = self.get_cell(self.this_row, self.this_col).value[0].id
        for c in model.get_dct_cells().values():
            if c.value[0].id == kat_id:
                self.SetCellBackgroundColour(c.row, c.col, WHITE)
                c.konzept = None
        self.ForceRefresh()

    def select_uncertain_from_popup(self, _):
        current_textcolor = frame.myGrid.GetCellTextColour(self.this_row, self.this_col)
        if active_konzeptColor != "#ffffff" and active_konzeptColor != current_textcolor:
            frame.myGrid.SetCellTextColour(self.this_row, self.this_col, active_konzeptColor)
            active_konzept.uncertain.append((self.this_row, self.this_col))
        else:
            frame.myGrid.SetCellTextColour(self.this_row, self.this_col, "#000000")
            active_konzept.uncertain.remove((self.this_row, self.this_col))
        self.ForceRefresh()

    def select_uncertain_cat_from_popup(self, _):
        current_cell = self.get_cell(self.this_row, self.this_col)
        current_textcolor = frame.myGrid.GetCellTextColour(self.this_row, self.this_col)
        for c in model.get_dct_cells().values():
            if c.value[0].id == current_cell.value[0].id:
                if active_konzeptColor != "#ffffff" and active_konzeptColor != current_textcolor:
                    frame.myGrid.SetCellTextColour(c.row, c.col, active_konzeptColor)
                    active_konzept.uncertain.append((c.row, c.col))
                else:
                    frame.myGrid.SetCellTextColour(c.row, c.col, "#000000")
                    active_konzept.uncertain.remove((c.row, c.col))
        self.ForceRefresh()

    def set_cellvalue(self, cellpos, value):
        cell_row, cell_col = cellpos
        self.SetCellValue(cell_row, cell_col, value)

    def select_cell(self, event):
        if active_konzeptColor != "#ffffff":    # active_konzeptColor == "#ffffff" würde bedeuten,
                                                # es ist kein Konzept ausgewählt. Mitlerweile wird beim erzeugen eines
                                                # Konzept dieses auch direkt auf Aktiv gesetzt und macht diese Bedingung
                                                # evtl überflüßig
            try:
                if model.RECHNUNGSTYP == "INST":
                    c = dct_cells[(event.GetRow(), event.GetCol())] # c = Zelle, die von der Maus angewählt wurde
                else:
                    c = model.get_dct_cells()[(self.this_row, self.this_col)]

                if model.RECHNUNGSTYP == "INST" and c.row < 9:
                    pass
                else:
                    if c.color != "#cccccc":
                        if c.color != active_konzeptColor:
                            c.color = active_konzeptColor
                            c.konzept = active_konzept
                            if (c.row, c.col) not in self.parent.GetParent().konzepte[c.konzept.name].cells:
                                self.parent.GetParent().konzepte[c.konzept.name].cells.append((c.row, c.col))
                        else:
                            if c.color != WHITE:
                                self.parent.GetParent().konzepte[c.konzept.name].cells.remove((c.row, c.col))
                            c.color = WHITE
                            c.konzept = None

                        self.SetCellBackgroundColour(c.row, c.col, c.color)

                        self.ForceRefresh()
                        self.parent.GetParent().tmpStoredKonzepte = None
                    else:
                        msg = "Grey cells are just for orientation,\nthey are not selectable."
                        dlg = wx.MessageDialog(self, msg, 'Information', style=wx.OK | wx.ICON_INFORMATION | wx.CENTRE)
                        dlg.ShowModal()
            except KeyError:
                pass
        else:
            msg = "Add or select a concept first."
            dlg = wx.MessageDialog(self, msg, 'Information', style=wx.OK | wx.ICON_INFORMATION | wx.CENTRE)
            dlg.ShowModal()

    def trigger_kategorie(self, event):
        if active_konzeptColor != "#ffffff":
            try:
                current_cell = self.get_cell(event.GetRow(), event.GetCol())
                kat_id = current_cell.value[0].id
                for c in model.get_dct_cells().values():
                    if c.value[0].id == kat_id:
                        if c.color != active_konzeptColor:
                            c.color = active_konzeptColor
                            c.konzept = active_konzept
                            if (c.row, c.col) not in self.parent.GetParent().konzepte[c.konzept.name].cells:
                                self.parent.GetParent().konzepte[c.konzept.name].cells.append((c.row, c.col))
                        else:
                            if c.color != WHITE:
                                self.parent.GetParent().konzepte[c.konzept.name].cells.remove((c.row, c.col))
                            c.color = WHITE
                            c.konzept = None
                        self.SetCellBackgroundColour(c.row, c.col, c.color)
                self.ForceRefresh()
            except KeyError:
                pass
        else:
            msg = "Add or select a concept first."
            dlg = wx.MessageDialog(self, msg, 'Information', style=wx.OK | wx.ICON_INFORMATION | wx.CENTRE)
            dlg.ShowModal()

    def reset_all_categories(self):
        tmp_dct = dct_cells.values() if model.RECHNUNGSTYP == "INST" else model.get_dct_cells().values()

        for c in tmp_dct:
            if self.GetCellBackgroundColour(c.row, c.col) != "#cccccc":
                self.SetCellBackgroundColour(c.row, c.col, WHITE)
                c.color = WHITE
            c.konzept = None
            frame.myGrid.SetCellTextColour(c.row, c.col, "#000000")

        self.ForceRefresh()
        self.parent.GetParent().tmpStoredKonzepte = None

    def delete_konzept_in_grid(self):
        if model.RECHNUNGSTYP == "INST":
            for c in dct_cells.values():
                if c.konzept == active_konzept:
                    self.SetCellBackgroundColour(c.row, c.col, WHITE)
                    c.konzept = None
                    c.color = WHITE
        else:
            for c in model.get_dct_cells().values():
                if c.konzept == active_konzept:
                    self.SetCellBackgroundColour(c.row, c.col, WHITE)
                    c.konzept = None
                    c.color = WHITE

        self.ForceRefresh()
        self.parent.GetParent().tmpStoredKonzepte = None

    def erase_grid(self):
        for row in range(200):
            for col in range(58):
                self.SetCellBackgroundColour(row, col, WHITE)
                self.set_cellvalue((row, col), "")

    def get_cell(self, row, col):
        return model.get_dct_cells()[(row, col)]


class InstitutsForm(wx.Frame):
    def __init__(self):
        """Constructor"""
        wx.Frame.__init__(self,
                          parent=None,
                          title="Finanzexplorer",
                          size=(1000, 700))
        self.p = wx.Panel(self, size=(300, 700))
        self.button = wx.Button(self.p, -1, label="plot me")
        self.search_control = wx.SearchCtrl(self.p, -1, value="", size=(240, 23))
        self.search_control.ShowCancelButton(True)
        self.btn_add = wx.Button(self.p, -1, label="add", size=(40, 20))
        self.btn_delete = wx.Button(self.p, -1, label="delete", size=(60, 20))
        self.btn_reset = wx.Button(self.p, -1, label="reset", size=(50, 20))
        self.btn_save = wx.Button(self.p, -1, label="save", size=(51, 20))
        self.btn_load = wx.Button(self.p, -1, label="load", size=(42, 20))

        self.static_text = wx.StaticText(self.p, -1, "Konzepte:")
        self.p.infotext = wx.StaticText(self.p, -1, "", size=(150, 23*4))
        # # self.p.infotext, um über parent vom grid (panel = self.p) auf den infotext zugreifen zu können
        self.list_ctrl_index = 0
        self.konzept_listcontrol = wx.ListCtrl(self.p, size=(155, 100),
                                               style=wx.LC_REPORT | wx.BORDER_SUNKEN | wx.LC_SINGLE_SEL)

        self.konzept_listcontrol.InsertColumn(0, 'Konzeptname', width=150)

        self.static_text_02 = wx.StaticText(self.p, -1, "Institutes:")
        self.checkbox_sum = wx.CheckBox(self.p, -1, label="Sum institutes")

        self.institute = defaultdict(dict)
        self.choices_files = [x for x in get_saves().keys()]
        self.choices = []

        self.checkbox = wx.CheckListBox(self.p, -1, size=(200, 1000),
                                        choices=[], style=wx.LB_ALWAYS_SB)
        self.checkbox.Bind(wx.EVT_CHECKLISTBOX, self.checked_item)

        self.myGrid = MyGrid(self.p)

        self.konzepte = {}
        self.tmpStoredKonzepte = None

        self.sizer_grid = wx.BoxSizer(wx.VERTICAL)
        self.sizer_grid.Add(self.search_control)
        self.sizer_grid.Add(self.myGrid, 1, wx.EXPAND)

        self.sizer_info = wx.BoxSizer(wx.VERTICAL)
        self.sizer_info.Add(self.p.infotext)

        self.sizer_save = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer_save.Add(self.btn_load)
        self.sizer_save.AddSpacer(60)
        self.sizer_save.Add(self.btn_save)
        # self.sizer_save.Add(self.filepicker)

        self.sizer_konzeptbutton = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer_konzeptbutton.AddSpacer(1)
        self.sizer_konzeptbutton.Add(self.btn_add)
        self.sizer_konzeptbutton.AddSpacer(1)
        self.sizer_konzeptbutton.Add(self.btn_delete)
        self.sizer_konzeptbutton.AddSpacer(1)
        self.sizer_konzeptbutton.Add(self.btn_reset)
        self.sizer_konzeptbutton.AddSpacer(1)

        self.sizer_instituts = wx.BoxSizer(wx.VERTICAL)
        self.sizer_instituts.Add(self.checkbox)

        self.sizer_controller = wx.BoxSizer(wx.VERTICAL)
        self.sizer_controller.AddSpacer(5)
        self.sizer_controller.Add(self.button)
        self.sizer_controller.AddSpacer(5)
        self.sizer_controller.Add(self.static_text)
        self.sizer_controller.AddSpacer(5)
        self.sizer_controller.Add(self.konzept_listcontrol)
        self.sizer_controller.Add(self.sizer_konzeptbutton)
        self.sizer_controller.AddSpacer(1)
        self.sizer_controller.Add(self.sizer_save)
        self.sizer_controller.AddSpacer(10)
        self.sizer_controller.Add(self.sizer_info)
        self.sizer_controller.Add(self.static_text_02)
        self.sizer_controller.Add(self.checkbox_sum)
        self.sizer_controller.Add(self.sizer_instituts)

        self.topSizer = wx.BoxSizer(wx.HORIZONTAL)
        self.topSizer.Add(self.sizer_controller)
        self.topSizer.Add(self.sizer_grid)

        self.p.SetSizer(self.topSizer)

        self.btn_load.Bind(wx.EVT_BUTTON, self.load_konzepte_picker)
        self.search_control.Bind(wx.EVT_SEARCHCTRL_SEARCH_BTN, self.find_category)
        self.search_control.Bind(wx.EVT_SEARCHCTRL_CANCEL_BTN, self.unfind_category)

        # --------- Menu Bar --------- #
        self.menuBar()

        self.Show()

    def menuBar(self):
        menuBar = wx.MenuBar()

        fileButton = wx.Menu()
        openMenu = wx.Menu()
        openItem01 = openMenu.Append(-1, 'Institutsexplorer')
        openItem02 = openMenu.Append(-1, 'Einnahmen-Ausgabenrechnung')
        openItem03 = openMenu.Append(-1, 'Vermögensübersicht')
        fileButton.Append(-1, 'Open ..', openMenu)

        openInNewWindowItem = fileButton.Append(-1, 'Open new Window')
        fileButton.AppendSeparator()
        create_report = fileButton.Append(-1, "create Report (PDF)")
        fileButton.AppendSeparator()
        exitItem = fileButton.Append(-1, 'Exit', 'status msg...')
        menuBar.Append(fileButton, 'File')

        runButton = wx.Menu()
        self.plotItem = runButton.Append(-1, 'create Graphs')
        menuBar.Append(runButton, 'Run')

        helpButton = wx.Menu()
        templateItem = helpButton.Append(-1, 'show Template')
        developeItem = helpButton.Append(-1, 'develope mode')
        menuBar.Append(helpButton, 'help')

        self.SetMenuBar(menuBar)
        self.Bind(wx.EVT_MENU, self.Quit, exitItem)
        self.Bind(wx.EVT_MENU, lambda event: self.set_interface_alert(event, 'INST'), openItem01)
        self.Bind(wx.EVT_MENU, lambda event: self.set_interface_alert(event, 'EA'), openItem02)
        self.Bind(wx.EVT_MENU, lambda event: self.set_interface_alert(event, 'VÜ'), openItem03)

        self.Bind(wx.EVT_MENU, self.new_window, openInNewWindowItem)
        self.Bind(wx.EVT_MENU, self.create_report, create_report)
        self.Bind(wx.EVT_MENU, self.open_template, templateItem)
        self.Bind(wx.EVT_MENU, self.develope_mode, developeItem)

    def develope_mode(self, _):
        """
        If develope_mode True, docstrings will be printed on console.
        """
        global develope_mode
        develope_mode = not develope_mode

    def Quit(self, _):
        self.Close()

    def open_template(self, _):
        """
        öffnet das Excel-Template, welches wir zum einen bei der Datenaufnahme aus den Haushaltsplänen
        als Vorlage verwendet haben, zum anderen wurde anhand des Templates das Grid-Interface für die
        Institute im Finanzexplorer erzeugt.

        :param _: wx-event (not used)
        """
        os.system('open Institute/Haushaltsbücher_MPI_Template.xlsx&')

    def new_window(self, _):
        """
        Öffnet den Finanzexplorer erneut, um zum Beispiel parallel an der Einnahmen-/Ausgabenrechnung und
        der Vermögensübersicht arbeiten zu können.

        :param _: wx-event (not used)
        """
        os.system('venv/bin/python Finanzexplorer_Institute.py&')

    def set_interface_alert(self, _, typ):
        """
        Falls bereits ein Interface eines bestimmten Rechnungstyps geöffnet ist und nicht dem entspricht,
        der gerade geöffnet werden soll, erscheint eine Warnung, dass noch ungespeicherte Konzepte verloren gehen.

        :param _: wx-event (not used)
        :param typ: Zur Unterscheidung von EA = Einnahmen-/Ausgabenrechnung, VÜ = Vermögensübersicht, INST = Institute
        :return:
        """
        if model.RECHNUNGSTYP == "":
            self.set_interface(_, typ)
        elif model.RECHNUNGSTYP == typ:
            pass
        else:
            msg = "Unsaved changes will be discarded!\n\nDo you want to proceed?"
            dlg = wx.MessageDialog(self, msg, 'Warning', style=wx.YES_NO | wx.ICON_QUESTION | wx.CENTRE)
            result = dlg.ShowModal()
            if result == wx.ID_YES:
                self.set_interface(_, typ)
            else:
                pass

    def set_interface(self, _, typ):
        """
        wxFrame wird nach dem 'richtigen' Rechnungstyp ausgerichtet: wxWidgets werden neue Funktionen zugeteilt.

        :param _: wx-event (not used)
        :param typ: Zur Unterscheidung von EA = Einnahmen-/Ausgabenrechnung, VÜ = Vermögensübersicht, INST = Institute
        :return:
        """

        # erase old interface and data
        fresh_new_start()
        frame.reset_konzepte(_)

        # set new interface and data
        model.RECHNUNGSTYP = typ

        if model.RECHNUNGSTYP == "EA":      # User-Interface for Einnahmen-/Ausgabenrechnung
            model.superkategorien = [68, 92]
            frame.SetTitle("Finanzexplorer - MPG Gesamt - Einnahmen-/ Ausgabenrechnung")
        elif model.RECHNUNGSTYP == "VÜ":    # User-Interface for Vermögensübersicht
            model.superkategorien = [1, 2]
            frame.SetTitle("Finanzexplorer - MPG Gesamt - Vermögensübersicht")

        # Superkategorien sind die 'Wurzeln' des Schema-Baums in EA und VÜ.
        # Also den Kategorien "Einnahmen", "Ausgaben", "Aktiva" und "Passiva".
        # Die Zahlen entsprechen den Kategorie-IDs in der Finanz-DB.

        if model.RECHNUNGSTYP == "INST":
            frame.SetTitle("Finanzexplorer - MPG Institute")

            self.checkbox_sum.SetLabel("Sum institutes")
            self.set_institutes()
            self.choices = [x for x in self.get_institutes().keys()]
            self.choices.sort()

            self.checkbox.Set(self.choices)
            import_inst_template()

            # Bindings for "INST"
            self.button.Bind(wx.EVT_BUTTON, self.button_click)
            self.myGrid.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.myGrid.select_cell)
            self.Bind(wx.EVT_MENU, self.button_click, self.plotItem)
            self.checked_item(_=None)
            self.sizer_controller.Show(self.checkbox_sum)
            self.sizer_controller.Layout()

        else:   # for "EA" or "VÜ"
            self.checkbox.Set([])

            self.choices_files = [x for x in get_saves().keys()]
            self.checkbox.Set(self.choices_files)

            import_mpg_gesamt_data()
            populate_cells()

            self.static_text_02.SetLabel("Saved Conzepts:")

            # Bindings for "EA" or "VÜ"
            self.button.Bind(wx.EVT_BUTTON, self.button_click_gesamt)
            self.myGrid.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.myGrid.trigger_kategorie)
            self.myGrid.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.myGrid.show_popup_menu)
            self.Bind(wx.EVT_MENU, self.button_click_gesamt, self.plotItem)

            self.sizer_controller.Hide(self.checkbox_sum)
            self.sizer_controller.Layout()

        # Bindings for all types
        self.btn_add.Bind(wx.EVT_BUTTON, self.add_konzept)
        self.konzept_listcontrol.Bind(wx.EVT_LIST_ITEM_SELECTED, self.colour_picked)
        self.konzept_listcontrol.Bind(wx.EVT_LIST_ITEM_DESELECTED, self.check_selection)
        self.btn_reset.Bind(wx.EVT_BUTTON, self.reset_konzepte)
        self.btn_delete.Bind(wx.EVT_BUTTON, self.delete_konzept)
        self.btn_save.Bind(wx.EVT_BUTTON, self.save_konzepte)

    def set_institutes(self):
        """
        Liest die Namen und Pfade aller relevanten Excel-Dateien aus dem Ordner 'Institute' ein und
        speichert sie im dict self.institute unter dem Key des Institutsnamens. Nested dict 'path' and 'file'
        :return:
        """
        if develope_mode:
            print(help(self.set_institutes))

        for (dirpath, dirnames, filenames) in os.walk("../Finanzexplorer-Git-data/Institute"):
            filenames = [x for x in filenames
                         if "Haushaltsb" in x
                         and ".xlsx" in x
                         and "Template" not in x
                         and "Haushaltsbücher_MPG_gesamt.xlsx" not in x
                         and "_All" not in x]
            for f in filenames:
                if f.split("_")[1] == "MPI":
                    name = f[21:len(f)-5].lower()
                else:
                    name = f[17:len(f) - 5].lower()
                name = name.capitalize()

                self.institute[name]["path"] = os.path.join(dirpath, f)
                self.institute[name]["file"] = f

    def get_institutes(self):
        return self.institute

    # Damit die Anzahl der ausgewählten Institute angezeigt wird
    def checked_item(self, _):
        if model.RECHNUNGSTYP == "INST":
            if len(self.checkbox.GetCheckedItems()) != 0:
                self.static_text_02.SetLabel("Institute ({} selected):".format(len(self.checkbox.GetCheckedItems())))
            else:
                self.static_text_02.SetLabel("Institute:")
        self.tmpStoredKonzepte = None

    def add_konzept(self, _):
        DialogNewKonzept(self, "new concept").ShowModal()
        # Das Fenster 'DialogNewKonzept' wird geöffnet. Name und Farbe des Konzepts können bestimmt werden
        # .ShowModal(), damit an anderen Fenstern nicht gearbeitet werden kann, solange dieses geöffnet ist
        # Die Alternative zu .ShowModal() wäre .Show()

    def add_new_konzept(self, name, color):
        self.konzept_listcontrol.InsertItem(self.list_ctrl_index, name)
        self.konzept_listcontrol.SetItemTextColour(self.list_ctrl_index, color)

        self.konzepte[name] = Konzept(name, color)
        self.konzept_listcontrol.SetItemState(self.list_ctrl_index, wx.LIST_STATE_SELECTED, wx.LIST_STATE_SELECTED)
        self.list_ctrl_index += 1
        self.tmpStoredKonzepte = None

    def check_selection(self, _):
        global active_konzeptColor
        global active_konzept
        if self.konzept_listcontrol.GetSelectedItemCount() == 0: # Wenn kein Konzept angewählt ist ..
            active_konzeptColor = "#ffffff"
            active_konzept = None

    def colour_picked(self, event):
        global active_konzeptColor
        global active_konzept
        active_konzeptColor = self.konzept_listcontrol.GetItemTextColour(event.GetIndex())
        active_konzept = self.konzepte[self.konzept_listcontrol.GetItem(event.GetIndex()).GetText()]

    def reset_konzepte(self, _):
        self.konzepte = {}
        self.konzept_listcontrol.DeleteAllItems()
        self.list_ctrl_index = 0
        self.myGrid.reset_all_categories()
        for item in self.checkbox.GetCheckedItems():
            self.checkbox.Check(item, False)
        self.checked_item(_=None)
        self.tmpStoredKonzepte = None

    def delete_konzept(self, _):
        global active_konzept, active_konzeptColor
        self.myGrid.delete_konzept_in_grid()

        # delete dct-entry
        self.tmpStoredKonzepte = None
        try:
            del self.konzepte[active_konzept.name]
        except AttributeError:
            pass

        # delete listCtrl-entry
        self.list_ctrl_index -= 1
        try:
            self.konzept_listcontrol.DeleteItem(self.konzept_listcontrol.GetFocusedItem())
        except AssertionError:
            pass

        # set new active konzept to last one in list...
        items_in_listctrl = self.konzept_listcontrol.GetItemCount()
        if items_in_listctrl > 0:
            active_konzept = self.konzepte[self.konzept_listcontrol.GetItem(items_in_listctrl - 1).GetText()]
            active_konzeptColor = active_konzept.color
            self.konzept_listcontrol.SetItemState(self.list_ctrl_index - 1,
                                                  wx.LIST_STATE_SELECTED,
                                                  wx.LIST_STATE_SELECTED)
        # ...or to "None", if list is empty
        else:
            active_konzept = None
            active_konzeptColor = WHITE

    def get_inst_konzepte(self):
        # print(get_inst_konzepte().__name__)
        if self.tmpStoredKonzepte:      # spart Zeit, da nicht erneut die Excel-Tabellen geöffnet werden müssen
            all_konzept_data = self.tmpStoredKonzepte
        else:
            # get_selected_institutes
            checked_institutes_paths = []
            for x in self.checkbox.GetCheckedItems():
                checked_institutes_paths.append((self.get_institutes()[self.choices[x]]["path"], self.choices[x]))

            # get existing concepts
            dct_xkonzepte = defaultdict(dict)
            tmp_xkonzepte = defaultdict(list)
            for c in dct_cells.values():
                if c.konzept:
                    tmp_xkonzepte[(worksheets[c.col], c.konzept.name)].append(c.value)

            for sheet_konzeptname, lst in tmp_xkonzepte.items():
                dct_xkonzepte[sheet_konzeptname[0]][sheet_konzeptname[1]] = lst
                # worksheets = ["1954-1963", "1964-1966", "1967", "1968-1972", "1973-1986", "1987-1997", "1998-2002"]
                # c.col = Spalte im interface = index aus der Liste 'worksheets'
                # dct_xkonzepte{"1954-1963": {Konzeptname:[list of selected cells per year]}}

            # get data according to the selected institutes AND existing concepts

            all_konzept_data = {}
            for path in checked_institutes_paths:
                konzept_data = defaultdict(list)
                wb = load_workbook(path[0], data_only=True)

                for sheet in worksheets:
                    try:
                        ws = wb[sheet]       # exp. Worksheet "1954-1963"

                        # get data (sheet by sheet) for each concept and store it in a list (each konzept has one list)
                        for konzeptname, conceptcells_per_year in dct_xkonzepte[sheet].items():
                            for row in range(1, ws.max_row+1):
                                for v in conceptcells_per_year:
                                    if ws.cell(row, 1).value == v:
                                        # hier wird in die Excel gegangen und
                                        # nach der ausgewählten Zelle aus dem Interface gesucht
                                        for col in range(1, ws.max_column+1):
                                            if ws.cell(3, col).value and ws.cell(row, col).value not in ["", " ", None,
                                                                                                         "None", "0", 0]:
                                                if sheet == "1998-2002" and col > 12:   # €-Werte
                                                    pass
                                                else:
                                                    konzept_data[konzeptname].append(
                                                        ExcelCell(row=row,
                                                                  col=col,
                                                                  path=path,
                                                                  sheet=sheet,
                                                                  year=int(ws.cell(3, col).value),
                                                                  typ=ws.cell(2, col).value,
                                                                  category=ws.cell(row, 1).value,
                                                                  betrag=float(ws.cell(row, col).value)))
                    except KeyError:
                        print("{} doesn't has sheet {}".format(path[0], sheet))
                    except ValueError:
                        print("\nWrong entry: {}\nPath: {}\nSheet: {}\nRow, Column: {}, {}\n".format(wb[sheet].cell(row, col).value, path[0], sheet, row, col))
                all_konzept_data[path[1]] = konzept_data

            for key, inst in all_konzept_data.items():
                for name, lst in inst.items():          # lst = list of ExcelCell objects
                    lst.sort(key=lambda z: z.year)
                    for obj in lst:
                        self.konzepte[name].plots[key].append((obj.year, obj.betrag, obj.type))

            # -----------------------------------
            # # Summe wird immer auch gespeichert. Ob sie genutzt wird entscheidet self.summe in der Klasse Konzepte
            for key, inst in all_konzept_data.items():
                for name, lst in inst.items():  # lst = list of ExcelCell objects; name = Konzeptname
                    for obj in lst:
                        tmp_is_in_list = False
                        for index, x in enumerate(self.konzepte[name].plots["SUMME"]):
                            if x[0] == obj.year and x[2] == obj.type:
                                tmp_lst = list(x)
                                tmp_lst[1] += obj.betrag
                                self.konzepte[name].plots["SUMME"][index] = tuple(tmp_lst)
                                tmp_is_in_list = True

                        if not tmp_is_in_list:
                            self.konzepte[name].plots["SUMME"].append((obj.year, obj.betrag, obj.type))
            # -----------------------------------

            self.tmpStoredKonzepte = all_konzept_data

        return all_konzept_data

    def button_click(self, _):
        if frame.checkbox_sum.GetValue():
            data = self.get_inst_konzepte()
            line_plot_inst(data, typ="IST", grouping_by="Institute", mode="auto", inflation=False)
        else:
            self.plot_settings()

    def button_click_gesamt(self, _):
        line_plot_gesamt_settings()

    def new_get_konzept(self):
        tmp_dct_xkonzepte = defaultdict(dict)
        dct_xkonzepte = defaultdict(list)

        for s in model.get_dct_schemata().values():
            if s.typ == model.RECHNUNGSTYP:
                for kon in self.konzepte.values():
                    tmp_dct_xkonzepte[kon.name][s.jahr] = None

                for c in s.cells:
                    if c.konzept:
                        if tmp_dct_xkonzepte[c.konzept.name][s.jahr] is None:
                            tmp_dct_xkonzepte[c.konzept.name][s.jahr] = 0.00
                        if c.jahr >= 2001:
                                tmp_dct_xkonzepte[c.konzept.name][s.jahr] += (c.posten.geldbetrag * 1.95583)   # DM -> €
                        else:
                            tmp_dct_xkonzepte[c.konzept.name][s.jahr] += c.posten.geldbetrag

        for name, konzept in tmp_dct_xkonzepte.items():
            for year, v_values in konzept.items():
                dct_xkonzepte[name].append((year, v_values))

            dct_xkonzepte[name].sort(key=lambda x: x[0])
            self.konzepte[name].plots["MPG-Gesamt"] = dct_xkonzepte[name]
        return dct_xkonzepte

    def plot_settings(self, xshow=True):
        data = self.get_inst_konzepte()
        return_code = str(DialogPlotSettings(self, title="plot settings").ShowModal())
        if len(return_code) == 5:   # Prüfnummer (Länge 5 bedeutet Fenster wurde nicht geschlossen)
            typ = ["IST", "SOLL", "IST & SOLL"][int(return_code[0]) - 1]
            grouping_by = ["Institute", "Konzept"][int(return_code[1]) - 1]
            mode = ["auto", "compareable"][int(return_code[2]) - 1]
            inflation = [False, True][int(return_code[3]) - 1]

            if not xshow:
                image = line_plot_inst(data, typ, grouping_by=grouping_by, mode=mode, inflation=inflation, xshow=False)
                return image
            else:
                line_plot_inst(data, typ, grouping_by=grouping_by, mode=mode, inflation=inflation, xshow=True)

        else:
            pass

    def save_konzepte(self, _):
        konzeptnames = []
        for x in self.checkbox.GetCheckedItems():
            konzeptnames.append(self.choices[x])
            konzeptnames.append("_")
        for k, v in self.konzepte.items():

            konzeptnames.append(k)
            konzeptnames.append("_")
        str_konzeptnames = ''.join(konzeptnames)
        str_konzeptnames = str_konzeptnames[:len(str_konzeptnames) - 1]     # um den letzen Unterstrich zu löschen
        str_konzeptnames = str_konzeptnames.replace("/", "-")               # um keinen subfolder zu erzeugen (Error)

        new_filename = model.RECHNUNGSTYP + "_" + str_konzeptnames

        filedlg = wx.FileDialog(self.p, "Save Konzept", defaultFile=new_filename, style=wx.FD_SAVE)
        filedlg.ShowModal()
        filepath = filedlg.GetPath()

        if filepath:
            if model.RECHNUNGSTYP == "INST":
                self.get_inst_konzepte()
            else:
                self.new_get_konzept()

            with open(filepath, "w") as outfile:
                json.dump(self.konzepte, outfile, default=lambda o: o.__dict__)

    def load_konzepte_picker(self, _):
        filedlg = wx.FileDialog(self.p, "Load", style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
        filedlg.ShowModal()
        filepath = filedlg.GetPath()
        if filepath:
            self.load_konzept(_, filepath)
            return True
        else:
            return False

    def load_konzept(self, _, filepath):
        global active_konzept, active_konzeptColor
        with open(filepath) as infile:
            dct_infile = json.load(infile)
            checked_items = []
            for k, v in dct_infile.items():     # k = Konzeptname, v = <dict> Kozept (Keys = Attr der Class Konzept)
                if model.RECHNUNGSTYP == "":
                    model.RECHNUNGSTYP = v["rechnungstyp"]
                    self.set_interface(_=None, typ=v["rechnungstyp"])
                if v["rechnungstyp"] == model.RECHNUNGSTYP:
                    self.add_new_konzept(k, v["color"])
                    for cell_pos in v["cells"]:
                        if model.RECHNUNGSTYP == "INST":
                            c = dct_cells[(cell_pos[0], cell_pos[1])]
                        else:
                            c = model.get_dct_cells()[(cell_pos[0], cell_pos[1])]
                        if c.color != v["color"]:
                            active_konzeptColor = v["color"]
                            c.color = active_konzeptColor
                            active_konzept = self.konzepte[k]
                            c.konzept = active_konzept
                            if (cell_pos[0], cell_pos[1]) not in self.konzepte[k].cells:
                                self.konzepte[k].cells.append((cell_pos[0], cell_pos[1]))
                        self.myGrid.SetCellBackgroundColour(c.row, c.col, c.color)
                    if v["rechnungstyp"] == "INST":
                        for inst in v["plots"].keys():
                            if inst != "SUMME":
                                checked_items.append(inst.capitalize())
                            else:
                                self.checkbox_sum.SetValue(True)
                    if v["rechnungstyp"] != "Inst":
                        for cell_pos in v["uncertain"]:
                            frame.myGrid.SetCellTextColour(cell_pos[0], cell_pos[1], v["color"])

                else:
                    msg = "File '{}' is not of type '{}'\n\n" \
                          "Unsaved changes will be discarded!\n\n" \
                          "Do you want to proceed?".format(filepath.split("/")[-1], model.RECHNUNGSTYP)
                    dlg = wx.MessageDialog(self, msg, 'Warning', style=wx.YES_NO | wx.ICON_QUESTION | wx.CENTRE)
                    result = dlg.ShowModal()
                    if result == wx.ID_YES:
                        model.RECHNUNGSTYP = v["rechnungstyp"]
                        self.set_interface(_, v["rechnungstyp"])
                        self.add_new_konzept(k, v["color"])
                        for cell_pos in v["cells"]:
                            if model.RECHNUNGSTYP == "INST":
                                c = dct_cells[(cell_pos[0], cell_pos[1])]
                            else:
                                c = model.get_dct_cells()[(cell_pos[0], cell_pos[1])]
                            if c.color != v["color"]:
                                active_konzeptColor = v["color"]
                                c.color = active_konzeptColor
                                active_konzept = self.konzepte[k]
                                c.konzept = active_konzept
                                if (cell_pos[0], cell_pos[1]) not in self.konzepte[k].cells:
                                    self.konzepte[k].cells.append((cell_pos[0], cell_pos[1]))
                            self.myGrid.SetCellBackgroundColour(c.row, c.col, c.color)
                        if v["rechnungstyp"] == "INST":
                            for inst in v["plots"].keys():
                                if inst != "SUMME":
                                    checked_items.append(inst.capitalize())
                    else:
                        break
            self.checkbox.SetCheckedStrings(checked_items)
            active_konzeptColor = "#ffffff"
            self.checked_item(_=None)
            self.myGrid.ForceRefresh()

    def find_category(self, _):
        self.unfind_category(_)

        value = self.search_control.GetValue()
        if model.RECHNUNGSTYP == "INST":
            for cell in dct_cells.values():
                if value.lower() in cell.value.lower() and cell.row > 8 and cell.color != "#cccccc":
                    found_cells.append(cell)
                    self.myGrid.SetCellBackgroundColour(cell.row, cell.col, "#d6ffff")
        else:
            for cell in model.get_dct_cells().values():
                if value.lower() in cell.value[0].bezeichnung.lower():
                    found_cells.append(cell)
                    self.myGrid.SetCellBackgroundColour(cell.row, cell.col, "#d6ffff")
        self.myGrid.ForceRefresh()

    def unfind_category(self, _):
        global found_cells
        for cell in found_cells:
            self.myGrid.SetCellBackgroundColour(cell.row, cell.col, cell.color)
        found_cells = []
        self.myGrid.ForceRefresh()

    def create_report(self, _, filename="Test"):
        if model.RECHNUNGSTYP == "INST":
            image = self.plot_settings(xshow=False)
        else:
            image = line_plot_gesamt(frame.new_get_konzept(), mode=0, xshow=False)

        myPDF = PDF()
        report_konzepte = defaultdict(dict)
        inhaltsverzeichnis = {}
        for konzeptname, konzeptobjekt in frame.konzepte.items():
            kategorien = []
            sorted_lst_cells = sorted(konzeptobjekt.cells, key=itemgetter(1))

            for konzeptcell in sorted_lst_cells:
                if model.RECHNUNGSTYP == "INST":
                    kategorien.append((dct_cells[konzeptcell].value, konzeptcell))
                else:
                    kategorien.append((model.get_dct_cells()[konzeptcell].value, konzeptcell, model.get_dct_cells()[konzeptcell].posten))
            if model.RECHNUNGSTYP == "INST":
                years = worksheets
            else:
                years = range(1948, 2006)

            for i, sheet in enumerate(years):
                xkategorien = []
                for kat in kategorien:
                    if kat[1][1] == i:
                        if model.RECHNUNGSTYP == "INST":
                            xkategorien.append(kat[0])
                        else:
                            if kat[0][0].spezifizierung:
                                xkategorien.append([kat[0][0].bezeichnung + " (" + kat[0][0].spezifizierung + "): ",
                                                   str(round(kat[2].geldbetrag, 2))])
                            else:
                                xkategorien.append([kat[0][0].bezeichnung + ": ", str(round(kat[2].geldbetrag, 2))])

                for item in xkategorien:
                    for i, x in enumerate(item):
                        item[i] = x.replace('–', '-')
                        item[i] = x.replace("\u0308", "_")

                report_konzepte[konzeptname][sheet] = xkategorien

        for toc in [False, True]:
            if toc:
                myPDF = PDF()
                myPDF.add_page()
                myPDF.set_font('Times', 'B', 12)
                myPDF.cell(w=200, h=12, txt="Inhaltsverzeichnis", ln=2)
                myPDF.set_font('Times', size=10)

                myPDF.cell(w=20, h=10, txt=" - ")
                link = myPDF.add_link()
                myPDF.set_link(link, page=2)
                myPDF.cell(w=100, h=10, txt="Diagramm", link=link, ln=1)

                for konzeptname, page in inhaltsverzeichnis.items():
                    myPDF.cell(w=20, h=10, txt=" - ")
                    link = myPDF.add_link()
                    myPDF.set_link(link, page=page + 1)
                    myPDF.cell(w=100, h=10, txt=konzeptname, link=link, ln=1)

            myPDF.set_font('Times', size=10)

            # Grafiken
            myPDF.add_page()
            if image:
                myPDF.image(image, w=200, h=150)

            for konzeptname, sheets in report_konzepte.items():
                myPDF.add_page()
                myPDF.set_font('Times', 'B', 12)
                myPDF.cell(w=50, h=12, txt="", ln=1)
                myPDF.cell(w=10, h=12, txt="")
                myPDF.cell(w=160, h=12, txt=konzeptname)
                myPDF.set_font('Times', size=8)
                myPDF.cell(w=50, h=12, txt="DM", ln=1)

                if not toc:
                    inhaltsverzeichnis[konzeptname] = myPDF.page_no()

                for sheet, txt in sheets.items():
                    x, y = myPDF.get_x(), myPDF.get_y()
                    myPDF.line(x + 20, y, x + 185, y)
                    summe = 0
                    for i, x in enumerate(txt):
                        myPDF.cell(w=10, h=8, txt="")
                        if i == 0:
                            myPDF.cell(w=10, h=8, txt=str(sheet))
                        else:
                            myPDF.cell(w=10, h=8, txt="")
                        y = myPDF.get_y()
                        myPDF.multi_cell(w=140, h=8, txt=x[0])
                        new_line_y = myPDF.get_y()
                        if new_line_y - y > 10:     # Wenn Multi_cell mit den Kategoriennamen einen linebreak hatte..
                            myPDF.set_xy(x=175, y=y+7)  # .. dann Cell mit Beträgen eine Zeile tiefer.
                        else:
                            myPDF.set_xy(x=175, y=y)    # .. sonst auf selber höhe.

                        myPDF.multi_cell(w=50, h=8, txt=x[1])
                        myPDF.set_y(new_line_y)
                        summe += float(x[1])
                    if len(txt) > 1:
                        x, y = myPDF.get_x(), myPDF.get_y()
                        myPDF.line(x + 165, y, x + 180, y)
                        myPDF.cell(w=10, h=8, txt="")
                        myPDF.cell(w=10, h=8, txt="")
                        myPDF.cell(w=145, h=8, txt="SUMME:")
                        myPDF.cell(w=50, h=8, txt=str(round(summe, 2)), ln=1)

            if toc:
                myPDF.output(filename + '.pdf', 'F')
                os.system('open Test.pdf&')
        myPDF.close()


def line_plot_inst(data, typ, grouping_by, mode, inflation, xshow=True):
    ymin, ymax, steps = get_limits(data, typ)
    fig = plt.figure()
    plt.close()
    if frame.checkbox_sum.GetValue():
        ax = {}
        for ex_institute in data.values():  # Bsp-Institut, um die Anzahl der Konzepte und Größe des Plot zu setzen
            if len(frame.konzepte.keys()) > 4:
                for count, konzeptname in enumerate(ex_institute.keys()):
                    rows = len(ex_institute.keys()) / 2
                    rows = int(rows) + 1 if rows % 1 != 0 else int(rows)
                    if count > rows - 1:
                        ax[konzeptname] = plt.subplot2grid((rows, 2), (count - rows, 1), rowspan=1, colspan=1)
                    else:
                        ax[konzeptname] = plt.subplot2grid((rows, 2), (count, 0), rowspan=1, colspan=1)
            else:
                count = 0
                for konzeptname in ex_institute.keys():
                    ax[konzeptname] = plt.subplot2grid((len(ex_institute.keys()) * 10, 1),
                                                       (count, 0), rowspan=8, colspan=1)
                    count += 10

        # dictionary mit den Werten 0 erzeugen
        dct_konzepte = {}
        for instname, konzepte in data.items():
            for konzeptname, ex_cells in konzepte.items():
                dct_konzepte[konzeptname] = {}
                for x in range(1954, 2003):
                    dct_konzepte[konzeptname][x] = 0

        # User fragen, ob der Plot Inflationsbereinigt sein soll oder nicht.
        bool_inflation = False
        msg = "Do you want this plot adjusted for inflation?"
        dlg = wx.MessageDialog(None, msg, 'Inflation', style=wx.YES_NO | wx.CANCEL | wx.CENTRE)
        result = dlg.ShowModal()
        if result == wx.ID_YES:
            bool_inflation = True
        elif result == 5101:
            return
        else:
            pass

        # auf jeden dict-wert den aktuellen Betrag draufaddieren
        dct_legend_set = defaultdict(set)
        for instname, konzepte in data.items():
            for konzeptname, ex_cells in konzepte.items():
                for x in ex_cells:
                    if x.type == typ:
                        if bool_inflation and x.year <= 1999:
                            dct_konzepte[konzeptname][x.year] += x.inflationsbereinigt
                        else:
                            dct_konzepte[konzeptname][x.year] += x.betrag
                        dct_legend_set[konzeptname].add(instname)

        for konzeptname, year in dct_konzepte.items():
            ax[konzeptname].plot(year.keys(), year.values(), color=frame.konzepte[konzeptname].color)
            ax[konzeptname].grid(linewidth='0.2')
            ax[konzeptname].ticklabel_format(style="plain")

        for konzeptname, instnames in dct_legend_set.items():
            tmp = [" + ".join(instnames), ]
            ax[konzeptname].legend(tmp)
            ax[konzeptname].set_title(konzeptname.upper(), pad=1)

    # ------------------------------
    else:
        if grouping_by == "Institute":
            ax = {}
            if len(frame.checkbox.GetCheckedItems()) > 4:
                for count, instname in enumerate(data.keys()):
                    rows = len(data.keys()) / 2
                    rows = int(rows) + 1 if rows % 1 != 0 else int(rows)

                    if count > rows - 1:
                        ax[instname] = plt.subplot2grid((rows, 2), (count - rows, 1), rowspan=1, colspan=1)
                    else:
                        ax[instname] = plt.subplot2grid((rows, 2), (count, 0), rowspan=1, colspan=1)
            else:
                count = 0
                for instname in data.keys():
                    ax[instname] = plt.subplot2grid((len(data.keys()) * 10, 1), (count, 0), rowspan=8, colspan=1)
                    count += 10

            for instname, inst in data.items():
                legend = []
                for key, kon in inst.items():
                    color = frame.konzepte[key].color
                    if typ == "IST & SOLL":
                        if inflation:
                            ax[instname].plot([x.year for x in kon if x.type == "IST" and x.year <= 1999],
                                              [y.inflationsbereinigt for y in kon if y.type == "IST" and y.year <= 1999],
                                              color=color)
                            legend.append(key + " (IST)")
                            ax[instname].plot([x.year for x in kon if x.type == "SOLL" and x.year <= 1999],
                                              [y.inflationsbereinigt for y in kon if y.type == "SOLL" and y.year <= 1999],
                                              color="#cccccc")
                            legend.append(key + " (SOLL)")
                        else:
                            ax[instname].plot([x.year for x in kon if x.type == "IST"],
                                              [y.betrag for y in kon if y.type == "IST"],
                                              color=color)
                            legend.append(key + " (IST)")
                            ax[instname].plot([x.year for x in kon if x.type == "SOLL"],
                                              [y.betrag for y in kon if y.type == "SOLL"],
                                              color="#cccccc")
                            legend.append(key + " (SOLL)")
                    else:
                        if inflation:
                            ax[instname].plot([x.year for x in kon if x.type == typ and x.year <= 1999],
                                              [y.inflationsbereinigt for y in kon if y.type == typ and y.year <= 1999],
                                              color=color)
                        else:
                            ax[instname].plot([x.year for x in kon if x.type == typ],
                                              [y.betrag for y in kon if y.type == typ],
                                              color=color)
                        legend.append(key)
                    ax[instname].set_title(instname.upper(), pad=1)

                    ax[instname].ticklabel_format(style="plain")
                    ax[instname].set_xticks([x for x in range(1945, 2010, 5)])
                    ax[instname].grid(linewidth='0.2')
                    if mode == "compareable":
                        ax[instname].set_yticks([y for y in range(ymin, ymax, steps)])

                ax[instname].legend(legend)

        elif grouping_by == "Konzept":
            ax = {}
            legend = defaultdict(list)
            for ex_institute in data.values():  # Bsp-Institut, um die Anzahl der Konzepte und Größe des Plot zu setzen
                if len(frame.konzepte.keys()) > 4:
                    for count, konzeptname in enumerate(ex_institute.keys()):
                        rows = len(ex_institute.keys()) / 2
                        rows = int(rows) + 1 if rows % 1 != 0 else int(rows)
                        if count > rows - 1:
                            ax[konzeptname] = plt.subplot2grid((rows, 2), (count - rows, 1), rowspan=1, colspan=1)
                        else:
                            ax[konzeptname] = plt.subplot2grid((rows, 2), (count, 0), rowspan=1, colspan=1)
                else:
                    count = 0
                    for konzeptname in ex_institute.keys():
                        ax[konzeptname] = plt.subplot2grid((len(ex_institute.keys()) * 10, 1),
                                                           (count, 0), rowspan=8, colspan=1)
                        count += 10

            for instname, inst in data.items():
                for key, kon in inst.items():
                    # color = frame.konzepte[key].color
                    if typ == "IST & SOLL":
                        if inflation:
                            ax[key].plot([x.year for x in kon if x.type == "IST" and x.year <= 1999],
                                         [y.inflationsbereinigt for y in kon if y.type == "IST" and y.year <= 1999])
                            legend[key].append(instname + " (IST)")
                            ax[key].plot([x.year for x in kon if x.type == "SOLL" and x.year <= 1999],
                                         [y.inflationsbereinigt for y in kon if y.type == "SOLL" and y.year <= 1999],
                                         color="#cccccc")
                            legend[key].append(instname + " (SOLL)")
                        else:
                            ax[key].plot([x.year for x in kon if x.type == "IST"],
                                         [y.betrag for y in kon if y.type == "IST"])
                            legend[key].append(instname + " (IST)")
                            ax[key].plot([x.year for x in kon if x.type == "SOLL"],
                                         [y.betrag for y in kon if y.type == "SOLL"],
                                         color="#cccccc")
                            legend[key].append(instname + " (SOLL)")
                    else:
                        if inflation:
                            ax[key].plot([x.year for x in kon if x.type == typ and x.year <= 1999],
                                         [y.inflationsbereinigt for y in kon if y.type == typ and y.year <= 1999])
                            legend[key].append(instname)
                        else:
                            ax[key].plot([x.year for x in kon if x.type == typ],
                                         [y.betrag for y in kon if y.type == typ])
                            legend[key].append(instname)
                    ax[key].set_title(key.upper(), pad=1)

                    ax[key].ticklabel_format(style="plain")
                    ax[key].set_xticks([x for x in range(1945, 2010, 5)])
                    ax[key].grid(linewidth='0.2')
                    if mode == "compareable":
                        ax[key].set_yticks([y for y in range(ymin, ymax, steps)])

            for key in legend.keys():
                ax[key].legend(legend[key])
# --------------------------
    if xshow:
        plt.show()
    else:
        plt.savefig('testchart.png', format='png', dpi=fig.dpi * 2)
        return 'testchart.png'


def line_plot_gesamt_settings(xshow=True):
    if len(frame.checkbox.GetCheckedItems()) != 0:
        checked_saves_paths = []
        for x in frame.checkbox.GetCheckedItems():
            checked_saves_paths.append([x for x in get_saves().values()][x])

        dlg = DialogGesamtPlotSettings(frame, "plot Settings", checked_saves_paths).ShowModal()
    else:
        line_plot_gesamt(frame.new_get_konzept(), mode=0, xshow=xshow)


def line_plot_gesamt(dct_konzepte, mode, xshow=True):
    legend = []
    fig = plt.figure()

    # User fragen, ob der Plot Inflationsbereinigt sein soll oder nicht.
    bool_inflation = False
    msg = "Do you want this plot adjusted for inflation?"
    dlg = wx.MessageDialog(None, msg, 'Inflation', style=wx.YES_NO | wx.CANCEL | wx.CENTRE)
    result = dlg.ShowModal()
    if result == wx.ID_YES:
        bool_inflation = True
    elif result == 5101:
        return
    else:
        pass

    # um "Nicht-Rechnungstypkonforme" Konzepte zu plotten------------------------------
    if len(frame.checkbox.GetCheckedItems()) != 0:
        checked_institutes_paths = []
        for x in frame.checkbox.GetCheckedItems():
            checked_institutes_paths.append([x for x in get_saves().values()][x])

        for path in checked_institutes_paths:
            with open(path) as infile:
                dct_infile = json.load(infile)
                for konzeptname, konzept_obj in dct_infile.items():
                    if len(konzept_obj["plots"].items()) == 1:
                        color = konzept_obj["color"]
                    else:
                        color = None

                    if mode == 1 and konzept_obj["rechnungstyp"] == "INST":
                        # mode 1: Die Summe der einzelnen Institute in einer gespeicherten Datei wird angezeigt.
                        x = []
                        y = []
                        konzept_obj["plots"]["SUMME"].sort(key=lambda x: x[0]) # nach Jahreszahlen (x[0]) sortieren
                        for data in konzept_obj["plots"]["SUMME"]:
                            if len(data) > 2:
                                if data[2] == "IST":
                                    if bool_inflation:
                                        if data[0] <= 1999:
                                            x.append(data[0])
                                            y.append(data[1] / dct_preisindices[data[0]])
                                    else:
                                        x.append(data[0])
                                        y.append(data[1])
                            else:
                                if bool_inflation:
                                    if data[0] <= 1999:
                                        x.append(data[0])
                                        y.append(data[1] / dct_preisindices[data[0]])
                                else:
                                    x.append(data[0])
                                    y.append(data[1])
                        plt.plot(x, y, color=color)
                        if (konzept_obj["rechnungstyp"] in ["EA", "VÜ"] and model.RECHNUNGSTYP in ["EA", "VÜ"]) or \
                                (konzept_obj["rechnungstyp"] == "INST" and model.RECHNUNGSTYP == "INST"):
                            legend.append(konzeptname)
                        else:
                            legend.append("SUMME - " + konzeptname)
                    elif mode == 2:
                        # mode2: Die einzelnen Institute aus den ausgewählten gespeicherten Dateien werden angezeigt.
                        for institute, values in konzept_obj["plots"].items():
                            if institute != "SUMME":
                                x = []
                                y = []
                                values.sort(key=lambda x: x[0]) # nach Jahreszahlen (x[0]) sortieren
                                for data in values:
                                    if len(data) > 2:
                                        if data[2] == "IST":
                                            if bool_inflation:
                                                if data[0] <= 1999:
                                                    x.append(data[0])
                                                    y.append(data[1] / dct_preisindices[data[0]])
                                            else:
                                                x.append(data[0])
                                                y.append(data[1])
                                    else:
                                        if bool_inflation:
                                            if data[0] <= 1999:
                                                x.append(data[0])
                                                y.append(data[1] / dct_preisindices[data[0]])
                                        else:
                                            x.append(data[0])
                                            y.append(data[1])

                                plt.plot(x, y, color=color)
                                if (konzept_obj["rechnungstyp"] in ["EA", "VÜ"] and model.RECHNUNGSTYP in ["EA", "VÜ"]) or \
                                        (konzept_obj["rechnungstyp"] == "INST" and model.RECHNUNGSTYP == "INST"):
                                    legend.append(konzeptname)
                                else:
                                    legend.append(institute + " - " + konzeptname)
                    elif mode == 3:
                        # mode 3: Summe und einzelne Institute werden angezeigt. Summe jeweils pro gespeicherte Datei!
                        for institute, values in konzept_obj["plots"].items():
                            x = []
                            y = []
                            values.sort(key=lambda x: x[0]) # nach Jahreszahlen (x[0]) sortieren
                            for data in values:
                                if len(data) > 2:
                                    if data[2] == "IST":
                                        if bool_inflation:
                                            if data[0] <= 1999:
                                                x.append(data[0])
                                                y.append(data[1] / dct_preisindices[data[0]])
                                        else:
                                            x.append(data[0])
                                            y.append(data[1])
                                else:
                                    if bool_inflation:
                                        if data[0] <= 1999:
                                            x.append(data[0])
                                            y.append(data[1] / dct_preisindices[data[0]])
                                    else:
                                        x.append(data[0])
                                        y.append(data[1])

                            plt.plot(x, y, color=color)
                            if (konzept_obj["rechnungstyp"] in ["EA", "VÜ"] and model.RECHNUNGSTYP in ["EA", "VÜ"]) or \
                                    (konzept_obj["rechnungstyp"] == "INST" and model.RECHNUNGSTYP == "INST"):
                                legend.append(konzeptname)
                            else:
                                legend.append(institute + " - " + konzeptname)
    # --------------------------------------------------------------------------------

    for key, lst in dct_konzepte.items():
        color = frame.konzepte[key].color
        values = []
        if bool_inflation:
            for x in lst:
                # prevent deviding a nonetype
                if x[1] is None and x[0] <= 1999:
                    values.append(None)
                elif x[0] <= 1999:
                    values.append(x[1] / dct_preisindices[x[0]] / 1000000)
            plt.plot([x[0] for x in lst if x[0] <= 1999], values, color=color)
        else:
            for x in lst:
                values.append(x[1] / 1000000)
            plt.plot([x[0] for x in lst], values, color=color)
        plt.grid(linewidth='0.2')
        legend.append(key)

    plt.ticklabel_format(style="plain")
    plt.xlim(1945, 2005)
    bottom, top = plt.ylim()
    plt.ylim(0, top)
    plt.legend(legend)
    plt.ylabel("Beträge in mio. DM")
    plt.xlabel("Jahre")
    if bool_inflation:
        plt.title("Inflationsbereinigt")
    else:
        plt.title("Title")
    if xshow:
        plt.show()
    else:
        plt.savefig('testchart.png', format='png', dpi=fig.dpi * 2)
        return 'testchart.png'


def import_inst_template():
    wb = load_workbook("../Finanzexplorer-Git-data/Institute/Haushaltsbücher_MPI_Template.xlsx", data_only=True)
    grid_col = 0
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "1954-1963":             # Sonderregel für dieses Sheet,
                                                        # da im Template die Einnahmen mehrmals auftauchen
            wb.active = s
            sheet = wb.active
            frame.myGrid.SetColLabelValue(grid_col, sheet.title)
            for row in range(1, 14):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)

            for row in range(56, 200):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)
            grid_col += 1
        elif wb.sheetnames[s][0] == "1":
            wb.active = s
            sheet = wb.active
            frame.myGrid.SetColLabelValue(grid_col, sheet.title)
            for row in range(1, 200):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
            grid_col += 1
    frame.myGrid.ForceRefresh()


tmp_stack = []


def set_hierarchie(schema_id, oberkategorie_id):
    if oberkategorie_id in model.superkategorien:
        model.get_dct_schemata()[schema_id].kategorien_hierarchisch.append(
            (model.get_dct_kategorien()[oberkategorie_id],
             len(tmp_stack) * "-",
             None))
    tmp_stack.append(oberkategorie_id)
    for kr in model.get_dct_schemata()[schema_id].KategorieRelationen:
        if model.get_dct_kategorien()[kr.origin_id].id == oberkategorie_id:
            model.get_dct_schemata()[schema_id].kategorien_hierarchisch.append(
                (kr.target_kategorie,
                 len(tmp_stack) * "-",
                 model.get_dct_kategorien()[kr.origin_id]))
            set_hierarchie(schema_id, kr.target_id)
    tmp_stack.pop()


def import_mpg_gesamt_data():
    model.import_kategorien()
    model.import_posten()
    model.import_kategorierelations()
    model.import_schemata()

    for year in range(1948, 2006):
        for schema in model.get_dct_schemata().values():
            if schema.jahr == year:
                for x in model.superkategorien:
                    set_hierarchie(schema.id, x)

    new_schemata = []
    last_schema = None
    for year in range(1948, 2006):
        has_schema = False
        for schema in model.get_dct_schemata().values():
            if schema.jahr == year:
                last_schema = copy.copy(schema)
                has_schema = True

        if not has_schema and last_schema:
            new_schema = copy.deepcopy(last_schema)
            new_schema.id += 1000
            new_schema.jahr = year
            new_schemata.append(new_schema)
            last_schema = new_schema

    for new in new_schemata:
        model.get_dct_schemata()[new.id] = new


def populate_cells():
    col = -1
    for year in range(1948, 2006):
        for schema in model.get_dct_schemata().values():
            if schema.jahr == year:
                row = 0
                col += 1

                frame.myGrid.SetColLabelValue(col, str(schema.jahr))
                for kategorie_in_Hierarchie in schema.kategorien_hierarchisch:
                    row += 1
                    temp = model.Cell(row, col, kategorie_in_Hierarchie, schema.jahr)
                    model.get_dct_cells()[(row, col)] = temp
                    model.get_dct_schemata()[schema.id].cells.add(temp)

    for cell in model.get_dct_cells().values():
        cell.calculate_zwischensumme()
        frame.myGrid.set_cellvalue(cell.get_pos(), str(cell.value[1]) + cell.value[0].bezeichnung)


def fresh_new_start():
    global dct_cells
    model.dct_Kategorien = {}
    model.dct_Posten = {}
    model.dct_categorierelations = OrderedDict()
    model.dct_Schemata = OrderedDict()
    model.dct_cells = {}

    dct_cells = {}
    frame.myGrid.erase_grid()


def get_saves():
    saves_path = {}
    for (dirpath, dirnames, filenames) in os.walk("../../../Saves"):
        filenames = [x for x in filenames]
        for f in filenames:
            filepath = os.path.join(dirpath, f)
            try:
                if "rechnungstyp" in open(filepath).read():
                    saves_path[f] = filepath
            except UnicodeDecodeError:
                pass
    return saves_path


def rgb_to_hex(rgb):
    r, g, b = rgb[:3]

    def clamp(x):
        return max(0, min(x, 255))
    return "#{0:02x}{1:02x}{2:02x}".format(clamp(r), clamp(g), clamp(b))


def get_limits(data, typ):
    tmp_min = 0
    tmp_max = 0
    for konzeptname, konzept in data.items():
        for institute, values in konzept.items():
            if typ == "IST & SOLL":
                for z in values:
                    if z.betrag > tmp_max:
                        tmp_max = z.betrag
                    if z.betrag < tmp_min:
                        tmp_min = z.betrag
            else:
                for z in values:
                    if z.betrag > tmp_max and z.type == typ:
                        tmp_max = z.betrag
                    if z.betrag < tmp_min and z.type == typ:
                        tmp_min = z.betrag

    tmp_min = int(tmp_min)
    tmp_max = int(tmp_max)
    tmp_range = tmp_max - tmp_min
    tmp_steps = int(math.pow(10, int(math.log10(tmp_range))-1))
    while tmp_min % tmp_steps:
        tmp_min -= 1
    while tmp_max % tmp_steps:
        tmp_max += 1

    tmp_steps *= 5

    return tmp_min, tmp_max, tmp_steps


def get_inflation_indices():
    dct_preisindices = {}

    wb = load_workbook("../Finanzexplorer-Git-data/verbraucherpreisindex-lange-reihen-xlsx-5611103.xlsx", data_only=True)
    ws = wb["JD_Index"]

    for row in range(6, 49):
        cell = ws.cell(row, 1).value
        dct_preisindices[int(cell[:4])] = float(ws.cell(row, 6).value / 100)

    for row in range(51, 60):
        cell = ws.cell(row, 1).value
        dct_preisindices[int(cell[:4])] = float(ws.cell(row, 6).value / 100)

    return dct_preisindices


'''def tracefunc(frame, event, arg, indent=[0]):
    if ("Finanzexplorer_model.py" in frame.f_code.co_filename or "Finanzexplorer_Institute.py" in frame.f_code.co_filename) \
            and "on_mouse_over" not in frame.f_code.co_name or "get_dct_cells" not in frame.f_code.co_name:
        if event == "call":
            indent[0] += 2
            print("-" * indent[0] + "> call function", frame.f_code.co_name)
        elif event == "return":
            print("<" + "-" * indent[0], "exit function", frame.f_code.co_name)
            indent[0] -= 2
    return tracefunc


import sys

sys.setprofile(tracefunc)'''



# create a directory 'Saves' next to the Platypus-file ( hopefully ;) )
os.makedirs("../Saves", exist_ok=True)

app = wx.App(False)
frame = InstitutsForm()

init_frame = DialogRechnungstypInit(None, "Start with ..")
init_frame.Destroy()

dct_preisindices = get_inflation_indices()

app.MainLoop()

print("Process finished with exit code 0")
