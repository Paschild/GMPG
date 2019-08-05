import openpyxl

wb = openpyxl.load_workbook(filename=r"/Users/mzichert/Documents/FE-Versuche/Haushaltsbücher_MPG_Test.xlsx")
x=1
for y in range(1, 20):
    c=wb.active.cell(row=y, column=x)
    print(getattr(c.fill.fgColor, c.fill.fgColor.type))




'''for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == "Tabelle1":
        wb.active = s
        i = s['A1'].fill.start_color.index  # Green Color
        Colors = styles.colors.COLOR_INDEX
        result = str(Coloros[i])'''