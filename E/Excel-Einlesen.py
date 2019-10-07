
from openpyxl import load_workbook

wb = load_workbook("/Users/easchadendorf/Documents/Mappe-PyCharm.xlsx")

wb.sheetnames

sheet = wb.active

print (sheet.cell(row=5, column=1).value)

print (sheet["A"])

print(sheet["A:B"])


'''for row in sheet.iter_rows(min_row=1,
                           max_row=3,
                           min_col=1,
                           max_col=3):

    print(row)'''  #Lee de manera horizontal (valor por row) la lista Excel




'''for col in sheet.iter_cols(min_col=1,
                           max_col=4,
                           min_row=5,
                           max_row=10,
                           values_only=True):


    print(col))'''   #Lee de manera vertical (valor por column) la lista Excel




'''for value in sheet.iter_rows(min_row=1,
                             max_row=3,
                             min_col=1,
                             max_col=4,
                             values_only=True):

    print(value)'''    #Da los valores internos de cada Cell


'''for column in sheet.columns:
    print(column)'''   #Da todos los valores de las columnas (o de los rows en su caso)



'''for value in sheet.iter_rows(min_row=2,
                             min_col=2,
                             max_col=4,
                             values_only=True):
    print(value)'''     #Igual que arriba
















