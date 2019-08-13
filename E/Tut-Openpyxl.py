from openpyxl import load_workbook

def read_excel(path):
    worksheets = ["1954-1963", "1964-1966", "1967", "1968-1972", "1973-1986", "1987-1997", "1998-2002"]
    wb = load_workbook(path, data_only=True)

    # --- für alle sheets
    for sheet in worksheets:
        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                print(ws.cell(row, col).value)


    # ---


# --- Main ---
read_excel(path="<Hier den richtigen Pfad zur Exceldatei eingeben>")


# --- Main End ---