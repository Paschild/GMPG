from openpyxl import load_workbook
import matplotlib.pyplot as plt
import wx
import wx.grid as gridlib
from collections import defaultdict

global istodersoll
istodersoll = 1
global istodersoll_text
istodersoll_text = ""


class MyGrid(wx.Frame):
    def __init__(self, new_dict_template):
        wx.Frame.__init__(self, parent = None, title = "Auswahlraster", size=(500, 250))
        panel = wx.Panel(self)




        myRaster = gridlib.Grid(panel)
        myRaster.CreateGrid(150, 7)
        myRaster.EnableEditing(False)

        '''Überschriften'''
        spalte = 0
        for key in new_dict_template:
            myRaster.SetCellValue(0, spalte, key)
            myRaster.SetCellFont(0, spalte, wx.Font(16, wx.ROMAN, wx.NORMAL, wx.BOLD))
            myRaster.SetCellBackgroundColour(0, spalte, wx.LIGHT_GREY)
            spalte = spalte + 1

        '''Zeilen füllen'''
        y = 0
        for key in new_dict_template:
            x = 2
            for value in new_dict_template[key]:
                myRaster.SetCellValue(x, y, value)
                myRaster.SetCellFont(x, y, wx.Font(14, wx.ROMAN, wx.NORMAL, wx.NORMAL))
                x = x + 1
            y = y + 1

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(myRaster)
        panel.SetSizer(sizer)


class AuswahlBox(wx.Frame):
    def __init__(self, parent, title):
        super(AuswahlBox, self).__init__(parent, title=title, size=(250, 200))

        panel = wx.Panel(self)
        box = wx.BoxSizer(wx.VERTICAL)
        auswahl = ["Auswahl treffen", "IST", "SOLL"]

        self.label = wx.StaticText(panel, label="IST oder SOLL?", style=wx.ALIGN_CENTER)
        box.Add(self.label, 0, wx.EXPAND | wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 15)

        self.combo = wx.ComboBox(panel, choices=auswahl)
        box.Add(self.combo, 0, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALL, 20)
        self.combo.Bind(wx.EVT_COMBOBOX, self.bei_Auswahl)

        self.button = wx.Button(panel, label="Okay", style=wx.ALIGN_CENTER_HORIZONTAL)
        box.Add(self.button, 0, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALL, 20)
        self.button.Bind(wx.EVT_BUTTON, self.bei_Click)


        panel.SetSizer(box)
        self.Center()
        self.Show()

    def bei_Auswahl(self, event):
        global istodersoll
        global istodersoll_text
        if self.combo.GetValue() == "IST":
            self.label.SetLabel("Es wurde " + self.combo.GetValue() + " ausgewählt")
            istodersoll = 1
            istodersoll_text = " (IST-Werte)"
        elif self.combo.GetValue() == "SOLL":
            self.label.SetLabel("Es wurde " + self.combo.GetValue() + " ausgewählt")
            istodersoll = 2
            istodersoll_text = " (SOLL-Werte)"
        else:
            self.label.SetLabel("Bitte IST oder SOLL auswählen")
            istodersoll = 0
            print("Bitte IST oder SOLL auswählen")

    def bei_Click(self, event):
        self.Close()


class Template_Cell:
    def __init__(self, row, column, value, color, sheet):
        self.row = row
        self.column = column
        self.value = value
        self.color = color
        self.sheet = sheet
        #self.konzepte = konzepte           #später kommt noch self.konzepte dazu


class Cell:
    def __init__(self, row, column, value, year, sheet, istodersoll):
        self.row = row
        self.column = column
        self.value = value
        self.year = year
        self.sheet = sheet
        if istodersoll == 1:
            self.istodersoll = "IST"
        elif istodersoll == 2:
            self.istodersoll = "SOLL"

    def ausgabe(self):
        print("Mappe: " + str(self.sheet))
        print("Jahr: " + str(self.year))
        print(str(self.istodersoll) + "-Wert")
        print("Spalte: " +str(self.column))
        print("Zeile: " + str(self.row))
        print("Wert: " + str(self.value))


def gettemplate():
    file_location = "/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPI_Template.xlsx"
    wb = load_workbook(file_location, data_only=True)
    template_cells = []
    dict_template = defaultdict(list)
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] != "Synthese":
            wb.active = sheet
            spalte = 1
            for zeile in range(10, 200):
                inhalt = wb.active.cell(row = zeile, column = spalte)
                if not inhalt.value:
                    continue
                else:
                    template_cells.append(Template_Cell(zeile, spalte, inhalt.value, getattr(inhalt.fill.fgColor, inhalt.fill.fgColor.type), wb.sheetnames[sheet]))
                    dict_template[wb.sheetnames[sheet]].append(inhalt.value)
    return template_cells, dict_template


def getdata(wanted):
    wb = load_workbook(filename, data_only=True)
    global istodersoll
    data = []
    years = []
    cells = []
    g = gen_years()
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] == "1954-1963":
            wb.active = sheet
            for zeile in range(1, 100):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 20, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column= istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1964-1966":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 4, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))
                    for spalte in range(4, 6, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1967":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 2, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1968-1972":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1973-1986":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 28, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1987-1997":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 22, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1998-2002":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

    return cells


def gen_years():
    for x in range(1954, 2003):
        yield x


def plot_matplot(new_cells):
    years = []
    data = []
    for cell in new_cells:
        years.append(cell.year)
        data.append(cell.value)

    plt.plot(years, data, "m")
    plt.title("Haushaltsplan" + istodersoll_text)
    plt.ylabel("Einnahmen/Ausgaben")
    plt.xlabel("Jahre")
    plt.text(60, .025, "Hallo")
    plt.show()


filename = "/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPG_Test.xlsx"
wanted = "Test"


#app = wx.App()
#AuswahlBox(None, 'Auswahlmenü')
#app.MainLoop()

#new_cells = getdata(wanted)
#plot_matplot(new_cells)

new_template_cells, new_dict_template = gettemplate()
#for cell in new_template_cells:
    #if cell.sheet == "1998-2002":
        #print(cell.value)


#for cell in new_cells:
    #cell.ausgabe()
    #break

app = wx.App()
frame = MyGrid(new_dict_template)
frame.Show()
app.MainLoop()




