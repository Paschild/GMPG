from openpyxl import load_workbook
import matplotlib.pyplot as plt
import wx

global istodersoll
istodersoll = 0
global istodersoll_text
istodersoll_text = ""

class AuswahlBox(wx.Frame):
    def __init__(self, parent, title):
        super(AuswahlBox, self).__init__(parent, title=title, size=(250, 200))

        panel = wx.Panel(self)
        box = wx.BoxSizer(wx.VERTICAL)
        auswahl = ["Auswahl treffen", "IST", "SOLL"]

        self.label = wx.StaticText(panel, label="IST oder SOLL?", style=wx.ALIGN_CENTER)
        box.Add(self.label, 0, wx.EXPAND | wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 15)

        self.combo = wx.ComboBox(panel, choices=auswahl)
        box.Add(self.combo, 0, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALL, 20)
        self.combo.Bind(wx.EVT_COMBOBOX, self.bei_Auswahl)

        self.button = wx.Button(panel, label="Okay", style=wx.ALIGN_CENTER_HORIZONTAL)
        box.Add(self.button, 0, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALL, 20)
        self.button.Bind(wx.EVT_BUTTON, self.bei_Click)


        panel.SetSizer(box)
        self.Center()
        self.Show()

    def bei_Auswahl(self, event):
        global istodersoll
        global istodersoll_text
        if self.combo.GetValue() == "IST":
            self.label.SetLabel("Es wurde " + self.combo.GetValue() + " ausgewählt")
            istodersoll = 1
            istodersoll_text = " (IST-Werte)"
        elif self.combo.GetValue() == "SOLL":
            self.label.SetLabel("Es wurde " + self.combo.GetValue() + " ausgewählt")
            istodersoll = 2
            istodersoll_text = " (SOLL-Werte)"
        else:
            self.label.SetLabel("Bitte IST oder SOLL auswählen")
            istodersoll = 0
            print("Bitte IST oder SOLL auswählen")

    def bei_Click(self, event):
        self.Close()


class Cell:
    def __init__(self, row, column, value, year, sheet, istodersoll):
        self.row = row
        self.column = column
        self.value = value
        self.year = year
        self.sheet = sheet
        if istodersoll == 1:
            self.istodersoll = "IST"
        elif istodersoll == 2:
            self.istodersoll = "SOLL"

    def ausgabe(self):
        print("Mappe: " + str(self.sheet))
        print("Jahr: " + str(self.year))
        print(str(self.istodersoll) + "-Wert")
        print("Spalte: " +str(self.column))
        print("Zeile: " + str(self.row))
        print("Wert: " + str(self.value))


def gen_years():
    for x in range(1954, 2003):
        yield x


def getdata(wanted):
    wb = load_workbook(filename, data_only=True)
    global istodersoll
    data = []
    years = []
    cells = []
    g = gen_years()
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] == "1954-1963":
            wb.active = sheet
            for zeile in range(1, 100):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 20, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column= istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1964-1966":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 4, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value/1000, year, wb.sheetnames[sheet], istodersoll))
                    for spalte in range(4, 6, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1967":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 2, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1968-1972":
            wb.active = sheet
            for zeile in range(1, 150):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1973-1986":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 28, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1987-1997":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 22, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

        elif wb.sheetnames[sheet] == "1998-2002":
            wb.active = sheet
            for zeile in range(1, 200):
                cell = wb.active.cell(row=zeile + 1, column=1)
                if cell.value == wanted:
                    for spalte in range(0, 10, 2):
                        inhalt = wb.active.cell(row=zeile + 1, column=istodersoll + spalte + 1)
                        data.append(inhalt.value)
                        year = next(g)
                        cells.append(Cell(zeile, spalte, inhalt.value, year, wb.sheetnames[sheet], istodersoll))

    return cells


def plot_matplot(new_cells):
    years = []
    data = []
    for cell in new_cells:
        years.append(cell.year)
        data.append(cell.value)

    plt.plot(years, data, "m")
    plt.title("Haushaltsplan" + istodersoll_text)
    plt.ylabel("Einnahmen/Ausgaben")
    plt.xlabel("Jahre")
    plt.text(60, .025, "Hallo")
    plt.show()


filename = "/Users/mzichert/Documents/FilesforFE/Haushaltsbücher_MPG_Test.xlsx"
wanted = "Test"


app = wx.App()
AuswahlBox(None, 'Auswahlmenü')
app.MainLoop()

new_cells = getdata(wanted)
plot_matplot(new_cells)

for cell in new_cells:
    cell.ausgabe()
    break
