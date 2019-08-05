"""Erkennt Farbe der Zellen und liest dann nur die gewünschten Zellen aus
mintgrün = FFCCFFCC
Braungrün-ähnliche Farbe = 6 (als int) """

from openpyxl import load_workbook
from collections import defaultdict


class Cell:
    def __init__(self, row, column, value, color, sheet):
        self.row = row
        self.column = column
        self.value = value
        self.color = color
        self.sheet = sheet

    def ausgabe(self):
        print(self.row)
        print(self.column)
        print(self.value)

def getdata():
    file_location = "Haushaltsbücher_MPG_Test.xlsx"
    wb = load_workbook(file_location, data_only=True)
    cells =     []
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "1954-1963":
            wb.active = s
            spalte = 1
            for y in range(1, 13):
                a1 = wb.active.cell(row=y, column=spalte)
                if a1.value == None:
                    continue
                if getattr(a1.fill.fgColor, a1.fill.fgColor.type) == "FFCCFFCC" or getattr(a1.fill.fgColor, a1.fill.fgColor.type) == 6:
                    cells.append(Cell(y, spalte, a1.value, getattr(a1.fill.fgColor, a1.fill.fgColor.type), wb.sheetnames[s]))

            for y in range(55, 70):
                a1 = wb.active.cell(row=y, column=spalte)
                if a1.value == None:
                    continue
                if getattr(a1.fill.fgColor, a1.fill.fgColor.type) == "FFCCFFCC" or getattr(a1.fill.fgColor, a1.fill.fgColor.type) == 6:
                    cells.append(Cell(y, spalte, a1.value, getattr(a1.fill.fgColor, a1.fill.fgColor.type), wb.sheetnames[s]))

        if wb.sheetnames[s] == "1964-1966":
            wb.active = s
            spalte = 1
            for y in range(1, 114):
                a1 = wb.active.cell(row=y, column=spalte)
                if a1.value == None:
                    continue
                if getattr(a1.fill.fgColor, a1.fill.fgColor.type) == "FFCCFFCC" or getattr(a1.fill.fgColor, a1.fill.fgColor.type) == 6:
                    cells.append(Cell(y, spalte, a1.value, getattr(a1.fill.fgColor, a1.fill.fgColor.type), wb.sheetnames[s]))

    return cells

new_cells = getdata()
liste_kategorien = []
dict =  defaultdict(list)
for cell in new_cells:
    #print("Zeile: "+ str(cell.row))
    #print("Inhalt: " + str(cell.value))
    #print("Mappe: " + str(cell.sheet) + "\n")
    dict[cell.sheet].append(cell.value)

    liste_kategorien.append(cell.value)

#print(liste_kategorien)
print(dict)



