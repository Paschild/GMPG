'''Anfänge bestimmte Zeilen finden und auslesen'''

from openpyxl import load_workbook
import matplotlib.pyplot as plt


def getdata(angabe):
    wb = load_workbook("Institute/Haushaltsbücher_MPG_Generalverwaltung.xlsx", data_only=True)
    data = []
    years = []
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "1954-1963":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row = y+1, column = 1)
                if a1.value == angabe:
                    for x in range(0, 20, 2):
                        b1 = wb.active.cell(row=y+1, column=x + 2)
                        data.append(b1.value)

        if wb.sheetnames[s] == "1964-1966":
            wb.active = s
            for y in range(1, 200):
                a1 = wb.active.cell(row = y+1, column = 1)
                if a1.value == angabe:
                    for x in range(0, 4, 2):
                        b1 = wb.active.cell(row=y+1, column=x + 2)
                        if b1.value == 0:
                            b1.value = None
                            data.append(b1.value)
                        else:
                            data.append(b1.value / 1000)
                    for x in range(4, 6, 2):
                        b1 = wb.active.cell(row=y + 1, column=x + 2)
                        if b1.value == 0:
                            b1.value = None
                        data.append(b1.value)


    for y in range(1954, 1967):
        years.append(y)

    print(data)

    plt.plot(years, data, "m")
    plt.title(wanted)
    plt.ylabel("Eingaben/Ausgaben")
    plt.xlabel("Jahre")
    plt.text(60, .025, "Hallo")
    plt.show()

wanted = "Gesamtausgaben"
getdata(wanted)

