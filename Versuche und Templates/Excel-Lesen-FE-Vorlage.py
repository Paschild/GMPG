'''Vorlage für lesen aus Excel-Dateien'''

from openpyxl import load_workbook

def import_inst_template():
    wb = load_workbook("Institute/Haushaltsbücher_MPI_Template.xlsx", data_only=True)
    grid_col = 0
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "1954-1963":             # Sonderregel für dieses Sheet,
                                                        # da im Template die Einnahmen mehrmals auftauchen
            wb.active = s
            sheet = wb.active
            frame.myGrid.SetColLabelValue(grid_col, sheet.title)
            for row in range(1, 14):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)

            for row in range(56, 200):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)
            grid_col += 1
        elif wb.sheetnames[s][0] == "1":
            wb.active = s
            sheet = wb.active
            frame.myGrid.SetColLabelValue(grid_col, sheet.title)
            for row in range(1, 200):
                if sheet.cell(row, 1).value:
                    this_cell = Cell(row, grid_col, sheet.cell(row, 1).value)
                    frame.myGrid.set_cellvalue((row, grid_col), sheet.cell(row, 1).value)
                    if sheet.cell(row, grid_col+1).fill.start_color.index == "00000000":
                        frame.myGrid.SetCellBackgroundColour(row, grid_col, "#cccccc")
                        this_cell.color = "#cccccc"
                    dct_cells[(row, grid_col)] = this_cell
            grid_col += 1
    frame.myGrid.ForceRefresh()


